"""
RMSecurity - Threat Intelligence de Ciberseguridad
Monitoreo de noticias de ciberseguridad con clasificación MITRE ATT&CK
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import feedparser
import requests
import re, os, sys, json, itertools
from datetime import date, datetime
from collections import defaultdict
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

def _resource(rel):
    base = getattr(sys, "_MEIPASS", OUTPUT_DIR)
    return os.path.join(base, rel)

# ─────────────────────────────────────────────────────────────
# SISTEMA DE TRADUCCIÓN (ES / EN)
# ─────────────────────────────────────────────────────────────
TRANSLATIONS = {
    "ES": {
        "title_top":       "◈ THREAT INTELLIGENCE · CIBERSEGURIDAD",
        "btn_search":      "◉ BUSCAR NOTICIAS",
        "btn_mark":        "✔ Marcar sel.",
        "btn_save_ioc":    "⬡ Guardar IoC",
        "btn_clear":       "↺ Limpiar",
        "btn_excel":       "▤ IoC Excel",
        "btn_feeds":       "▦ Feeds",
        "btn_open_link":   "↗ Abrir en navegador",
        "lbl_category":    "Categoría:",
        "lbl_severity":    "Severidad:",
        "lbl_search":      "Buscar:",
        "all":             "TODAS",
        "col_date":        "Fecha ↓",
        "col_check":       "✔",
        "col_sev":         "Severidad",
        "col_cat":         "Categoría",
        "col_mitre":       "MITRE ATT&CK",
        "col_title":       "Título",
        "col_actor":       "Actor / Grupo",
        "col_source":      "Fuente",
        "panel_analysis":  "▸ ANÁLISIS DE AMENAZA",
        "panel_ioc":       "IoC / INDICADORES",
        "panel_summary":   "RESUMEN",
        "loading_title":   "⟳  BUSCANDO AMENAZAS",
        "loading_feeds":   "Consultando feeds de ciberseguridad...",
        "status_ready":    "Sistema listo.  Presiona  ◉ BUSCAR  para comenzar.",
        "status_loading":  "Buscando...",
        "ioc_actor":       "Actor/Grupo",
        "ioc_origin":      "Origen",
        "ioc_victim":      "Víctima",
        "ioc_sector":      "Sector",
        "ioc_country_v":   "País víctima",
        "ioc_impact":      "Impacto",
        "ioc_data":        "Datos robados",
        "ioc_systems":     "Sistemas afect.",
        "ioc_ransom":      "Rescate/Monto",
        "ioc_software":    "Software",
        "ioc_versions":    "Versiones",
        "ioc_techniques":  "Técnicas T1",
        "ioc_ips":         "IPs C2/Mal.",
        "ioc_ports":       "Puertos",
        "ioc_domains":     "Dominios mal.",
        "ioc_hashes":      "Hashes",
        "ioc_wallets":     "Wallets crypto",
        "ioc_emails":      "Emails",
        "lang_btn":        "🌐 EN",
        "threats_found":   "amenazas detectadas",
        "search_done":     "Búsqueda completada",
        "no_results":      "Sin resultados para",
        "saving_ioc":      "⟳ Guardando IoC de",
        "saving_ioc2":     "noticias...",
        "saved_ioc":       "✔ IoC guardados:",
        "saved_new":       "nuevos registros →",
        "cleared":         "Limpiado.",
        "no_marked":       "No hay noticias marcadas.",
        "sev_values":      ["TODAS", "CRÍTICO", "ALTO", "MEDIO", "BAJO"],
    },
    "EN": {
        "title_top":       "◈ THREAT INTELLIGENCE · CYBERSECURITY",
        "btn_search":      "◉ SEARCH NEWS",
        "btn_mark":        "✔ Mark sel.",
        "btn_save_ioc":    "⬡ Save IoC",
        "btn_clear":       "↺ Clear",
        "btn_excel":       "▤ IoC Excel",
        "btn_feeds":       "▦ Feeds",
        "btn_open_link":   "↗ Open in browser",
        "lbl_category":    "Category:",
        "lbl_severity":    "Severity:",
        "lbl_search":      "Search:",
        "all":             "ALL",
        "col_date":        "Date ↓",
        "col_check":       "✔",
        "col_sev":         "Severity",
        "col_cat":         "Category",
        "col_mitre":       "MITRE ATT&CK",
        "col_title":       "Title",
        "col_actor":       "Actor / Group",
        "col_source":      "Source",
        "panel_analysis":  "▸ THREAT ANALYSIS",
        "panel_ioc":       "IoC / INDICATORS",
        "panel_summary":   "SUMMARY",
        "loading_title":   "⟳  SEARCHING THREATS",
        "loading_feeds":   "Querying cybersecurity feeds...",
        "status_ready":    "System ready.  Press  ◉ SEARCH  to begin.",
        "status_loading":  "Searching...",
        "ioc_actor":       "Actor/Group",
        "ioc_origin":      "Origin",
        "ioc_victim":      "Victim",
        "ioc_sector":      "Sector",
        "ioc_country_v":   "Victim country",
        "ioc_impact":      "Impact",
        "ioc_data":        "Stolen data",
        "ioc_systems":     "Aff. systems",
        "ioc_ransom":      "Ransom/Amount",
        "ioc_software":    "Software",
        "ioc_versions":    "Versions",
        "ioc_techniques":  "T1 Techniques",
        "ioc_ips":         "C2/Mal. IPs",
        "ioc_ports":       "Ports",
        "ioc_domains":     "Mal. domains",
        "ioc_hashes":      "Hashes",
        "ioc_wallets":     "Crypto wallets",
        "ioc_emails":      "Emails",
        "lang_btn":        "🌐 ES",
        "threats_found":   "threats detected",
        "search_done":     "Search completed",
        "no_results":      "No results for",
        "saving_ioc":      "⟳ Saving IoC from",
        "saving_ioc2":     "articles...",
        "saved_ioc":       "✔ IoC saved:",
        "saved_new":       "new records →",
        "cleared":         "Cleared.",
        "no_marked":       "No marked articles.",
        "sev_values":      ["ALL", "CRITICAL", "HIGH", "MEDIUM", "LOW"],
    },
}

_LANG = "ES"

def T(key):
    return TRANSLATIONS[_LANG].get(key, key)

# ─────────────────────────────────────────────────────────────
# PALETA NEON FUTURISTA
# ─────────────────────────────────────────────────────────────
C = {
    "bg0":      "#03050D",   # fondo principal (negro azulado)
    "bg1":      "#070C18",   # fondo secundario
    "bg2":      "#0B1120",   # fondo paneles
    "bg3":      "#0F1830",   # fondo filas alternas
    "border":   "#1A2A45",   # bordes sutiles
    "txt":      "#8899BB",   # texto base
    "txt_hi":   "#C8D8F0",   # texto destacado
    "cyan":     "#00F0FF",   # cyan neon — headers, títulos
    "green":    "#39FF14",   # verde neon — BAJO / OK
    "yellow":   "#FFE600",   # amarillo neon — MEDIO
    "orange":   "#FF7800",   # naranja neon — ALTO
    "red":      "#FF1040",   # rojo neon — CRÍTICO
    "violet":   "#C000FF",   # violeta neon — APT
    "pink":     "#FF0090",   # magenta neon — INFRA CRÍTICA
    "blue":     "#0090FF",   # azul neon — FILTRACIÓN
    "amber":    "#FFAA00",   # ámbar — PHISHING
    "lime":     "#A0FF20",   # lima — DARKWEB
    "sel":      "#0A2040",   # fila seleccionada
}

# Categorías: (bg_fila, fg_texto)  — fondo muy oscuro, texto neon
CAT_COLORS = {
    "RANSOMWARE":                   ("#1C0008", C["red"]),
    "VULNERABILIDADES":             ("#140A00", C["orange"]),
    "APT / ESPIONAJE":              ("#0E0018", C["violet"]),
    "FILTRACIÓN DE DATOS":          ("#00101A", C["blue"]),
    "MALWARE / TROYANOS":           ("#180010", C["pink"]),
    "INFRAESTRUCTURA CRÍTICA":      ("#1A0010", "#FF3080"),
    "PHISHING / INGENIERÍA SOCIAL": ("#130900", C["amber"]),
    "CIBERCRIMEN / DARKWEB":        ("#060F06", C["lime"]),
}

SEVERITY_COLORS = {
    "CRÍTICO": (C["red"],    "#1A0006"),
    "ALTO":    (C["orange"], "#140800"),
    "MEDIO":   (C["yellow"], "#14110000"[:-2]),
    "BAJO":    (C["green"],  "#051405"),
}

# ─────────────────────────────────────────────────────────────
# MITRE ATT&CK
# ─────────────────────────────────────────────────────────────
MITRE_TACTICS = {
    "TA0001": "Initial Access",
    "TA0002": "Execution",
    "TA0003": "Persistence",
    "TA0004": "Privilege Escalation",
    "TA0005": "Defense Evasion",
    "TA0006": "Credential Access",
    "TA0007": "Discovery",
    "TA0008": "Lateral Movement",
    "TA0009": "Collection",
    "TA0010": "Exfiltration",
    "TA0011": "C2 / Command & Control",
    "TA0040": "Impact",
    "TA0042": "Resource Development",
    "TA0043": "Reconnaissance",
}

MITRE_MAP = {
    "TA0001": ["phishing", "spear-phishing", "spear phishing", "drive-by", "supply chain",
               "cadena de suministro", "watering hole", "valid accounts", "acceso inicial",
               "brute force login", "credential stuffing", "password spray"],
    "TA0002": ["macro", "powershell", "script malicioso", "code execution", "ejecución de código",
               "shellcode", "command injection", "remote code execution", "rce",
               "ejecución remota", "wmi execution", "scheduled task execution"],
    "TA0003": ["persistencia", "persistence", "backdoor", "rootkit", "scheduled task",
               "tarea programada", "registry", "startup", "autorun", "boot kit",
               "servicio malicioso", "implant"],
    "TA0004": ["escalada de privilegios", "privilege escalation", "uac bypass",
               "token impersonation", "kernel exploit", "lpe", "dll hijacking",
               "sudo abuse", "setuid"],
    "TA0005": ["evasión", "obfuscation", "antivirus bypass", "sandbox evasion",
               "log tampering", "anti-forensic", "packed malware", "lolbins",
               "living off the land", "timestomping", "process injection"],
    "TA0006": ["credential dumping", "mimikatz", "pass the hash", "kerberoasting",
               "golden ticket", "lsass", "contraseñas robadas", "credential stuffing",
               "hash cracking", "password theft", "robo de credenciales"],
    "TA0007": ["network scan", "nmap", "enumeración", "enumeration", "active directory recon",
               "port scan", "escaneo de puertos", "asset discovery"],
    "TA0008": ["lateral movement", "pivoting", "rdp abuse", "psexec", "wmi lateral",
               "smb", "pass-the-ticket", "remote services", "ssh tunneling", "movimiento lateral"],
    "TA0009": ["keylogger", "screen capture", "captura de pantalla", "clipboard",
               "email collection", "data collection", "recolección de datos"],
    "TA0010": ["exfiltración", "exfiltration", "data leak", "fuga de datos", "data theft",
               "robo de datos", "dns exfiltration", "c2 exfil"],
    "TA0011": ["command and control", "c2", "c&c", "botnet", "beacon", "cobalt strike",
               "cobaltstrike", "metasploit", "meterpreter", "reverse shell",
               "domain fronting", "comunicación c2"],
    "TA0040": ["ransomware", "cifrado de archivos", "wiper", "destrucción de datos",
               "ddos", "denial of service", "denegación de servicio", "defacement",
               "sabotaje", "disruptivo", "datos cifrados"],
    "TA0042": ["infraestructura maliciosa", "bulletproof hosting", "dominio malicioso",
               "staging server", "exploit kit", "malicious infrastructure"],
    "TA0043": ["reconocimiento", "reconnaissance", "osint", "escaneo masivo",
               "fingerprinting", "shodan", "censys", "recon", "information gathering"],
}

# Técnicas específicas T1xxx
MITRE_TECHNIQUES = {
    "T1566": "Phishing", "T1190": "Exploit Public-Facing App",
    "T1195": "Supply Chain Compromise", "T1133": "External Remote Services",
    "T1059": "Command and Scripting Interpreter", "T1055": "Process Injection",
    "T1078": "Valid Accounts", "T1053": "Scheduled Task/Job",
    "T1003": "OS Credential Dumping", "T1110": "Brute Force",
    "T1486": "Data Encrypted for Impact", "T1491": "Defacement",
    "T1498": "Network Denial of Service", "T1071": "App Layer Protocol C2",
    "T1041": "Exfiltration Over C2 Channel", "T1567": "Exfiltration Over Web Service",
}

# ─────────────────────────────────────────────────────────────
# CATEGORÍAS
# ─────────────────────────────────────────────────────────────
CATEGORIAS = {
    "RANSOMWARE": [
        "ransomware", "extorsión digital", "rescate digital", "cifrado de archivos",
        "lockbit", "conti", "blackcat", "alphv", "clop", "hive", "rhysida",
        "play ransomware", "akira", "medusa", "royal ransomware", "double extortion",
        "doble extorsión", "pago en cripto", "descifrador", "decryptor", "ransom note",
        "datos cifrados", "bitcoin ransom", "datos secuestrados",
    ],
    "VULNERABILIDADES": [
        "vulnerabilidad", "vulnerability", "cve-", "zero-day", "0-day", "parche",
        "patch tuesday", "exploit", "poc", "proof of concept", "buffer overflow",
        "sql injection", "xss", "cross-site", "csrf", "ssrf", "rce", "lpe",
        "out-of-bounds", "use-after-free", "race condition", "path traversal",
        "log4j", "log4shell", "shellshock", "heartbleed", "eternalblue",
        "critical patch", "security advisory", "nvd", "cvss",
    ],
    "APT / ESPIONAJE": [
        "apt", "advanced persistent threat", "estado-nación", "nation-state",
        "espionaje cibernético", "cyber espionage", "lazarus", "fancy bear",
        "cozy bear", "apt28", "apt29", "apt41", "volt typhoon", "salt typhoon",
        "sandworm", "turla", "kimsuky", "mustang panda", "earth preta",
        "campaña de espionaje", "spyware", "pegasus", "predator", "intellexa",
        "cyber espionage campaign", "state-sponsored", "espionaje estatal",
    ],
    "FILTRACIÓN DE DATOS": [
        "filtración", "data breach", "brecha de datos", "datos expuestos",
        "leak", "fuga de información", "database exposed", "base de datos expuesta",
        "haveibeenpwned", "datos personales robados", "información sensible",
        "pii exposed", "credentials leaked", "contraseñas filtradas",
        "millions of records", "millones de registros", "datos vendidos",
        "darkweb sale", "venta datos", "exposed database",
    ],
    "MALWARE / TROYANOS": [
        "malware", "troyano", "trojan", "infostealer", "stealer", " rat ",
        "remote access trojan", "botnet", "worm", "gusano", "virus informático",
        "backdoor", "rootkit", "dropper", "loader", "emotet", "qakbot",
        "agent tesla", "redline", "raccoon", "formbook", "remcos",
        "cobalt strike", "brute ratel", "malware campaign", "campaña de malware",
    ],
    "INFRAESTRUCTURA CRÍTICA": [
        "infraestructura crítica", "critical infrastructure", "ics", "scada",
        "ot security", "industrial control", "power grid", "red eléctrica",
        "hospital hackeado", "healthcare breach", "water treatment", "planta de agua",
        "energy sector", "sector energético", "oleoducto", "pipeline hack",
        "sistema de control industrial", "ataque a hospital", "nuclear",
    ],
    "PHISHING / INGENIERÍA SOCIAL": [
        "phishing", "spear phishing", "smishing", "vishing", "whaling",
        "ingeniería social", "social engineering", "correo fraudulento",
        "estafa digital", "fraude online", "qr phishing", "quishing",
        "deepfake", "suplantación de identidad", "impersonation",
        "business email compromise", "bec", "fraude ceo",
    ],
    "CIBERCRIMEN / DARKWEB": [
        "darkweb", "dark web", "darknet", "tor network", "foro criminal",
        "hacker forum", "ransomhub", "ciberdelincuente", "cybercriminal",
        "money mule", "lavado criptográfico", "crypto laundering",
        "carding", "skimming digital", "fraude bancario online",
        "fbi seized", "interpol cyber", "operación policial cyber",
        "detenido por hackeo", "arrestado hacker", "breach forum",
        "raidforums", "xss forum", "exploit forum",
    ],
}

CATEGORIAS_ORDEN = list(CATEGORIAS.keys())

EXCLUIR = [
    "videojuego", "minecraft", "fortnite", "moda", "receta", "cocina",
    "fútbol", "deporte", "farándula", "horóscopo", "clima", "turismo",
    "entertainment", "celebrity", "gaming review", "movie", "película",
]

# ─────────────────────────────────────────────────────────────
# FEEDS RSS
# ─────────────────────────────────────────────────────────────
FEEDS = [
    ("Krebs on Security",       "https://krebsonsecurity.com/feed/"),
    ("BleepingComputer",        "https://www.bleepingcomputer.com/feed/"),
    ("The Hacker News",         "https://feeds.feedburner.com/TheHackersNews"),
    ("Dark Reading",            "https://www.darkreading.com/rss.xml"),
    ("SecurityWeek",            "https://feeds.feedburner.com/securityweek"),
    ("Recorded Future / TheRecord", "https://therecord.media/feed"),
    ("CyberScoop",              "https://cyberscoop.com/feed/"),
    ("SC Magazine",             "https://www.scmagazine.com/feed"),
    ("Infosecurity Magazine",   "https://www.infosecurity-magazine.com/rss/news/"),
    ("Help Net Security",       "https://www.helpnetsecurity.com/feed/"),
    ("Naked Security Sophos",   "https://nakedsecurity.sophos.com/feed/"),
    ("Graham Cluley",           "https://grahamcluley.com/feed/"),
    ("Schneier on Security",    "https://www.schneier.com/feed/atom/"),
    ("CISA Alerts",             "https://www.cisa.gov/news.xml"),
    ("CISA Advisories",         "https://www.cisa.gov/cybersecurity-advisories/feed"),
    ("US-CERT",                 "https://www.cisa.gov/uscert/ncas/alerts.xml"),
    ("ENISA",                   "https://www.enisa.europa.eu/news/enisa-news/RSS"),
    ("INCIBE España",           "https://www.incibe.es/rss.xml"),
    ("Mandiant Blog",           "https://www.mandiant.com/resources/blog/rss.xml"),
    ("CrowdStrike Blog",        "https://www.crowdstrike.com/blog/feed/"),
    ("SentinelOne Blog",        "https://www.sentinelone.com/blog/feed/"),
    ("Palo Alto Unit42",        "https://unit42.paloaltonetworks.com/feed/"),
    ("Talos Intelligence",      "https://blog.talosintelligence.com/feeds/posts/default"),
    ("ESET WeLiveSecurity",     "https://www.welivesecurity.com/en/feed/"),
    ("Kaspersky Securelist",    "https://securelist.com/feed/"),
    ("Microsoft Security",      "https://www.microsoft.com/en-us/security/blog/feed/"),
    ("Google Project Zero",     "https://googleprojectzero.blogspot.com/feeds/posts/default"),
    ("Google TAG",              "https://blog.google/threat-analysis-group/rss/"),
    ("NCC Group Research",      "https://research.nccgroup.com/feed/"),
    ("Rapid7 Blog",             "https://www.rapid7.com/blog/feed/"),
    ("Exploit-DB",              "https://www.exploit-db.com/rss.xml"),
    ("Segu.info",               "https://www.segu-info.com.ar/rss.php"),
    ("DragonJAR Colombia",      "https://www.dragonjar.org/feed"),
    ("Hackplayers España",      "https://www.hackplayers.com/feeds/posts/default"),
    ("GNews Ransomware AR",
     "https://news.google.com/rss/search?q=ransomware+Argentina&hl=es-419&gl=AR&ceid=AR:es-419"),
    ("GNews Ciberataque AR",
     "https://news.google.com/rss/search?q=ciberataque+Argentina&hl=es-419&gl=AR&ceid=AR:es-419"),
    ("GNews Hackeo AR",
     "https://news.google.com/rss/search?q=hackeo+filtración+datos+Argentina&hl=es-419&gl=AR&ceid=AR:es-419"),
    ("GNews APT ES",
     "https://news.google.com/rss/search?q=APT+ciberespionaje+estado-nación&hl=es-419&gl=ES&ceid=ES:es"),
    ("GNews Vuln ES",
     "https://news.google.com/rss/search?q=vulnerabilidad+critica+parche+CVE&hl=es-419&gl=ES&ceid=ES:es"),
    ("GNews Ransomware EN",
     "https://news.google.com/rss/search?q=ransomware+attack+2025&hl=en-US&gl=US&ceid=US:en"),
    ("GNews Data breach EN",
     "https://news.google.com/rss/search?q=data+breach+exposed+2025&hl=en-US&gl=US&ceid=US:en"),
    ("GNews Zero-day EN",
     "https://news.google.com/rss/search?q=zero-day+exploit+critical+2025&hl=en-US&gl=US&ceid=US:en"),
    ("GNews APT nation-state",
     "https://news.google.com/rss/search?q=APT+nation-state+cyberattack+espionage&hl=en-US&gl=US&ceid=US:en"),
    ("GNews CISA advisory",
     "https://news.google.com/rss/search?q=CISA+advisory+critical+vulnerability&hl=en-US&gl=US&ceid=US:en"),
    ("GNews Darkweb arrest",
     "https://news.google.com/rss/search?q=darkweb+cybercrime+arrested+seized&hl=en-US&gl=US&ceid=US:en"),
    ("GNews Critical infra",
     "https://news.google.com/rss/search?q=critical+infrastructure+cyberattack+SCADA&hl=en-US&gl=US&ceid=US:en"),
]

# ─────────────────────────────────────────────────────────────
# OLLAMA
# ─────────────────────────────────────────────────────────────
_OLLAMA_URL   = "http://localhost:11434/api/generate"
_OLLAMA_MODEL = "llama3.2:3b"
_OLLAMA_OK    = None
_UMBRAL_DUP   = 0.72
_stats_path   = os.path.join(OUTPUT_DIR, "feed_stats.json")
_ioc_path     = os.path.join(OUTPUT_DIR, "THREAT_INTEL_IOC.xlsx")

def ollama_disponible():
    global _OLLAMA_OK
    if _OLLAMA_OK is not None:
        return _OLLAMA_OK
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=2)
        _OLLAMA_OK = r.status_code == 200
    except Exception:
        _OLLAMA_OK = False
    return _OLLAMA_OK

def ollama_analizar_cyber(titulo, resumen):
    prompt = (
        "Eres un analista de ciberseguridad experto. Analiza esta noticia y responde SOLO con JSON válido.\n"
        f"Título: {titulo}\nTexto: {resumen[:600]}\n\n"
        "Devuelve exactamente este JSON:\n"
        '{"categoria": "<RANSOMWARE|VULNERABILIDADES|APT / ESPIONAJE|FILTRACIÓN DE DATOS|'
        'MALWARE / TROYANOS|INFRAESTRUCTURA CRÍTICA|PHISHING / INGENIERÍA SOCIAL|'
        'CIBERCRIMEN / DARKWEB|IRRELEVANTE>", '
        '"resumen_es": "<resumen claro en español de 2-3 oraciones con los datos clave>", '
        '"actor": "<nombre del grupo/actor amenaza o null>", '
        '"victima": "<organización/sector víctima o null>", '
        '"pais_victima": "<país de la víctima o null>", '
        '"mitre_tactica": "<ID táctica MITRE TA00xx más relevante o null>", '
        '"cve": "<número CVE si se menciona o null>", '
        '"impacto": "<descripción del impacto en 1 frase o null>"}'
    )
    try:
        r = requests.post(_OLLAMA_URL, json={
            "model": _OLLAMA_MODEL, "prompt": prompt, "stream": False,
            "options": {"temperature": 0, "num_predict": 300}
        }, timeout=12)
        txt = r.json().get("response", "")
        m = re.search(r'\{.*\}', txt, re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception:
        pass
    return None

# ─────────────────────────────────────────────────────────────
# LIMPIEZA DE TEXTO
# ─────────────────────────────────────────────────────────────
_RE_HTML  = re.compile(r'<[^>]+>')
_RE_ENT   = re.compile(r'&(?:amp|lt|gt|quot|nbsp|apos|#\d+);')
_ENT_MAP  = {"&amp;": "&", "&lt;": "<", "&gt;": ">", "&quot;": '"',
             "&nbsp;": " ", "&apos;": "'"}

def limpiar_texto(txt):
    if not txt:
        return ""
    txt = _RE_HTML.sub(" ", txt)
    for ent, rep in _ENT_MAP.items():
        txt = txt.replace(ent, rep)
    txt = _RE_ENT.sub("", txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt

def condensar_resumen(titulo, resumen):
    resumen = limpiar_texto(resumen)
    titulo  = limpiar_texto(titulo)
    if len(resumen) < 60:
        return resumen or titulo
    partes = re.split(r'(?<=[.!?])\s+', resumen)
    buenas = [p for p in partes if len(p) > 30][:3]
    return " ".join(buenas)[:600] if buenas else resumen[:600]

# ─────────────────────────────────────────────────────────────
# DESCARGA DE ARTÍCULO COMPLETO
# ─────────────────────────────────────────────────────────────
_HEADERS_FETCH = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
}
# dominios que bloquean scraping o redirigen — no intentar
_NO_FETCH = {"google.com", "news.google.com", "t.co", "twitter.com",
             "facebook.com", "linkedin.com", "paywall.com"}

def _fetch_articulo(url, timeout=4):
    """Descarga el texto plano del artículo completo. Devuelve str o ''."""
    if not url:
        return ""
    try:
        dom = urlparse(url).netloc.replace("www.", "")
        if any(nd in dom for nd in _NO_FETCH):
            return ""
        r = requests.get(url, headers=_HEADERS_FETCH, timeout=timeout,
                         allow_redirects=True, stream=False)
        if r.status_code != 200:
            return ""
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(r.text, "html.parser")
            for tag in soup(["script","style","nav","footer","header","aside"]):
                tag.decompose()
            art = soup.find("article") or soup.find("main") or soup.body
            return limpiar_texto(art.get_text(" ", strip=True)) if art else ""
        except ImportError:
            return limpiar_texto(_RE_HTML.sub(" ", r.text))
    except Exception:
        return ""


def _fetch_feed(url, timeout=5):
    """Descarga un feed RSS con timeout explícito y lo parsea."""
    try:
        r = requests.get(url, headers=_HEADERS_FETCH, timeout=timeout,
                         allow_redirects=True)
        return feedparser.parse(r.content)
    except Exception:
        try:
            return feedparser.parse(url)
        except Exception:
            return feedparser.FeedParserDict(entries=[])

# ─────────────────────────────────────────────────────────────
# EXTRACCIÓN IoC COMPLETA
# ─────────────────────────────────────────────────────────────
_RE_CVE      = re.compile(r'CVE-\d{4}-\d{4,7}', re.I)
_RE_IP       = re.compile(r'\b(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}'
                           r'(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\b')
_RE_HASH_MD5 = re.compile(r'\b[0-9a-fA-F]{32}\b')
_RE_HASH_SHA = re.compile(r'\b[0-9a-fA-F]{40,64}\b')
_RE_EMAIL    = re.compile(r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b')
_RE_BTC      = re.compile(r'\b(?:[13][a-km-zA-HJ-NP-Z1-9]{25,34}|bc1[a-z0-9]{39,59})\b')
_RE_ETH      = re.compile(r'\b0x[0-9a-fA-F]{40}\b')
_RE_MONERO   = re.compile(r'\b4[0-9AB][1-9A-HJ-NP-Za-km-z]{93}\b')
_RE_CVSS     = re.compile(r'CVSS(?:\s*v\d)?[\s:=]+(\d+\.?\d*)', re.I)
_RE_TECNICA  = re.compile(r'\bT1\d{3}(?:\.\d{3})?\b')
_RE_PORT     = re.compile(r'\bport[s]?\s+(\d{1,5})\b', re.I)
_RE_VERSION  = re.compile(
    r'(?:version|versión|v)\s*(\d+(?:\.\d+){1,3})', re.I)
_RE_DOMAIN_MAL = re.compile(
    r'\b(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)'
    r'+(?:ru|cn|ir|kp|su|xyz|top|tk|ml|ga|cf|gq|pw|cc|biz|onion)\b', re.I)

# Montos de dinero / rescates
_RE_DINERO = re.compile(
    r'(?:'
    r'\$\s*[\d,]+(?:\.\d+)?\s*(?:million|billion|thousand|M|B|K)?'
    r'|[\d,]+(?:\.\d+)?\s*(?:million|billion|thousand)\s*(?:dollar|euro|pound|USD|EUR)?'
    r'|\d+(?:\.\d+)?\s*BTC'
    r'|\d+(?:\.\d+)?\s*(?:XMR|ETH|monero)'
    r')',
    re.I
)
_RE_DATOS_ROBADOS = re.compile(
    r'(?:'
    r'[\d,.]+\s*(?:million|billion|thousand|GB|TB|MB)?\s*'
    r'(?:records?|users?|customers?|accounts?|patients?|employees?|'
    r'registros?|usuarios?|clientes?|pacientes?|empleados?|datos|archivos?|files?)'
    r')',
    re.I
)
_RE_SISTEMAS_AFECT = re.compile(
    r'(?:[\d,.]+\s*(?:systems?|devices?|computers?|servers?|machines?|'
    r'equipos?|servidores?|sistemas?)(?:\s+(?:infected|encrypted|compromised|affected))?)',
    re.I
)

GRUPOS_APT = [
    # Grupos estado-nación
    "lazarus group", "lazarus", "hidden cobra", "guardians of peace",
    "fancy bear", "apt28", "sofacy", "strontium", "forest blizzard",
    "cozy bear", "apt29", "midnight blizzard", "nobelium", "the dukes",
    "apt41", "winnti", "barium", "double dragon",
    "sandworm", "voodoo bear", "electrum", "seashell blizzard",
    "turla", "snake", "venomous bear", "waterbug",
    "kimsuky", "thallium", "velvet chollima",
    "mustang panda", "earth preta", "bronze president",
    "volt typhoon", "bronze silhouette",
    "salt typhoon", "ghostemperor", "fearful bug",
    "charming kitten", "phosphorus", "apt35", "mint sandstorm",
    "muddywater", "seedworm", "temp.zagros",
    "gamaredon", "armageddon", "primitive bear",
    "scattered spider", "unc3944", "octo tempest",
    # Ransomware grupos
    "lockbit", "lockbit 3.0", "lockbit 2.0",
    "alphv", "blackcat",
    "clop", "ta505",
    "conti", "wizard spider",
    "rhysida", "akira", "play ransomware", "medusa",
    "royal", "hive", "revil", "sodinokibi",
    "darkside", "blackmatter",
    "evil corp", "indrik spider",
    "lapsus$", "dev-0537",
    "fin7", "carbanak", "navigator group",
    "fin8", "sangria tempest",
    "silence group", "unc2452", "unc2891",
    "black basta", "qakbot", "qbot",
    "cuba ransomware", "industrial spy",
    "8base", "hunters international", "inc ransom",
    "ransomhub", "noname057",
]

SECTORES = {
    "hospital": "Salud", "healthcare": "Salud", "medical": "Salud",
    "health": "Salud", "clínica": "Salud", "clinic": "Salud",
    "pharma": "Farmacéutica", "pharmaceutical": "Farmacéutica",
    "bank": "Finanzas", "financial": "Finanzas", "banco": "Finanzas",
    "insurance": "Seguros", "fintech": "Finanzas",
    "energy": "Energía", "oil": "Energía", "gas ": "Energía",
    "power grid": "Energía", "electricidad": "Energía", "nuclear": "Energía",
    "government": "Gobierno", "gobierno": "Gobierno", "ministerio": "Gobierno",
    "federal": "Gobierno", "municipal": "Gobierno", "military": "Defensa",
    "defense": "Defensa", "nato": "Defensa/OTAN", "ejército": "Defensa",
    "university": "Educación", "universidad": "Educación", "school": "Educación",
    "education": "Educación", "college": "Educación",
    "telecom": "Telecomunicaciones", "isp": "Telecomunicaciones",
    "retail": "Comercio minorista", "ecommerce": "E-commerce",
    "airline": "Aviación", "airport": "Aviación", "aviation": "Aviación",
    "transport": "Transporte", "logistics": "Logística",
    "manufacturing": "Manufactura", "factory": "Manufactura",
    "water": "Agua/Servicios", "wastewater": "Agua/Servicios",
    "law firm": "Legal", "legal": "Legal",
    "media": "Medios", "newspaper": "Medios", "broadcasting": "Medios",
}

SOFTWARE_AFECTADO = [
    # OS / Plataformas
    "windows", "linux", "macos", "android", "ios",
    "windows server", "windows 10", "windows 11",
    # Microsoft
    "exchange", "sharepoint", "outlook", "teams", "office 365",
    "azure", "active directory", "hyper-v", "mshtml",
    # Network
    "cisco", "fortinet", "fortigate", "palo alto", "juniper",
    "netscaler", "citrix", "vmware", "esxi", "vcenter",
    "ivanti", "pulse secure", "globalprotect",
    # Web / App
    "apache", "nginx", "iis", "tomcat", "wordpress",
    "drupal", "joomla", "log4j", "spring", "struts",
    # Databases
    "mysql", "mssql", "postgresql", "mongodb", "redis",
    "elasticsearch", "oracle database",
    # Cloud
    "aws", "azure", "gcp", "s3 bucket", "kubernetes", "docker",
    # Security
    "vpn", "rdp", "ssh", "smb", "winrm",
]

PAISES = {
    "russia": "Rusia", "rusia": "Rusia", "russian": "Rusia",
    "china": "China", "chinese": "China",
    "iran": "Irán", "iranian": "Irán",
    "north korea": "Corea del Norte", "corea del norte": "Corea del Norte",
    "dprk": "Corea del Norte", "north korean": "Corea del Norte",
    "united states": "EEUU", "estados unidos": "EEUU", "american": "EEUU",
    "argentina": "Argentina",
    "brazil": "Brasil", "brasil": "Brasil",
    "mexico": "México", "méxico": "México",
    "spain": "España", "españa": "España",
    "germany": "Alemania", "alemania": "Alemania",
    "united kingdom": "Reino Unido", "reino unido": "Reino Unido",
    "ukraine": "Ucrania", "ucrania": "Ucrania",
    "israel": "Israel",
    "india": "India", "pakistan": "Pakistán",
    "taiwan": "Taiwán", "japan": "Japón",
    "australia": "Australia", "canada": "Canadá",
    "france": "Francia", "italia": "Italia", "italy": "Italia",
    "colombia": "Colombia", "chile": "Chile", "peru": "Perú",
}

def extraer_ioc(titulo, resumen, url="", texto_completo=""):
    """
    Extrae IoC del título + resumen + (opcionalmente) artículo completo.
    texto_completo debe ser el texto plano del artículo ya descargado.
    """
    # Combinar fuentes de texto: artículo completo > resumen
    base  = limpiar_texto(titulo + " " + resumen)
    full  = texto_completo if texto_completo else base
    # Para patrones técnicos usar texto completo; para contexto usar base
    texto_l = full.lower()
    base_l  = base.lower()

    ioc = {
        # ── Indicadores técnicos ──
        "cve":          sorted(set(_RE_CVE.findall(full)), key=lambda x: x)[:8],
        "ips":          [ip for ip in set(_RE_IP.findall(full))
                         if not ip.startswith(("192.168.","10.","172.1",
                                               "172.2","172.3","127.","0.","169."))
                         ][:8],
        "hashes_md5":   list(set(_RE_HASH_MD5.findall(full)))[:4],
        "hashes_sha":   list(set(_RE_HASH_SHA.findall(full)))[:4],
        "emails":       [e for e in set(_RE_EMAIL.findall(full))
                         if not e.endswith((".png",".jpg",".gif",".css"))][:5],
        "btc_wallets":  list(set(_RE_BTC.findall(full)))[:3],
        "eth_wallets":  list(set(_RE_ETH.findall(full)))[:3],
        "xmr_wallets":  list(set(_RE_MONERO.findall(full)))[:2],
        "dominios_mal": list(set(_RE_DOMAIN_MAL.findall(full)))[:6],
        "tecnicas":     list(set(_RE_TECNICA.findall(full)))[:6],
        "puertos":      list(set(_RE_PORT.findall(full)))[:5],
        # ── Contexto ──
        "cvss_score":       None,
        "grupos":           [],
        "victima":          None,
        "sector_victima":   None,
        "software_afectado":[],
        "versiones":        [],
        "pais_victima":     None,
        "pais_origen":      None,
        "rescate":          None,
        "datos_robados":    None,
        "sistemas_afectados": None,
        "impacto_resumen":  None,
    }

    # CVSS score
    m = _RE_CVSS.search(full)
    if m:
        ioc["cvss_score"] = m.group(1)

    # Rescate / dinero
    montos = _RE_DINERO.findall(full)
    if montos:
        # filtrar el más relevante (el más largo suele ser más específico)
        ioc["rescate"] = max(montos, key=len).strip()[:80]

    # Datos robados / sistemas afectados
    dr = _RE_DATOS_ROBADOS.findall(full)
    if dr:
        ioc["datos_robados"] = max(dr, key=len).strip()[:80]
    sa = _RE_SISTEMAS_AFECT.findall(full)
    if sa:
        ioc["sistemas_afectados"] = max(sa, key=len).strip()[:80]

    # Grupos APT / ransomware
    for g in GRUPOS_APT:
        if g in texto_l:
            canonical = g.title()
            if canonical not in ioc["grupos"]:
                ioc["grupos"].append(canonical)

    # Sector víctima
    for kw, sector in SECTORES.items():
        if kw in texto_l and not ioc["sector_victima"]:
            ioc["sector_victima"] = sector

    # Software afectado
    sw_encontrado = []
    for sw in SOFTWARE_AFECTADO:
        if sw in texto_l and sw.title() not in sw_encontrado:
            sw_encontrado.append(sw.title())
    ioc["software_afectado"] = sw_encontrado[:6]

    # Versiones de software
    vers = _RE_VERSION.findall(full)
    ioc["versiones"] = list(set(vers))[:4]

    # Víctima (múltiples patrones, de más a menos específico)
    patrones_victima = [
        # inglés — organización nombrada explícitamente
        r'(?:attacked|breached|hacked|ransomed|targeted|compromised|hit)\s+([A-Z][a-zA-Z0-9\s&\-\'\.]{2,40}?)(?:\s+(?:said|confirmed|reported|announced|told)|[,.])',
        r'([A-Z][a-zA-Z0-9\s&\-\'\.]{3,35}?)\s+(?:was hacked|suffered a(?:n)? (?:breach|attack|incident)|confirmed a breach|paid ransom)',
        r'(?:attack on|breach at|incident at|ransomware hit)\s+([A-Z][a-zA-Z0-9\s&\-\'\.]{3,35})',
        # español
        r'(?:atacó a|ataca a|víctima|comprometió a|afecta a|afectó a)\s+([A-ZÁÉÍÓÚa-záéíóú0-9\s&\-]{3,40}?)(?:\s+(?:dijo|confirmó|reveló)|[,.])',
        r'([A-ZÁÉÍÓÚa-z][a-záéíóú0-9\s&\-]{3,35}?)\s+(?:fue hackeada?|sufrió un ataque|pagó rescate)',
        r'(?:ataque a|brecha en|incidente en)\s+([A-ZÁÉÍÓÚa-z][a-záéíóú0-9\s&\-]{3,35})',
    ]
    for pat in patrones_victima:
        m = re.search(pat, full)
        if m:
            candidato = m.group(1).strip()
            # descartar si parece una palabra genérica o muy corta
            if len(candidato) > 3 and not candidato.lower() in (
                "the", "a", "an", "its", "their", "company", "organization",
                "sector", "government", "la", "el", "un", "una"
            ):
                ioc["victima"] = candidato[:55]
                break

    # País víctima (toma el primero mencionado)
    for kw, pais in PAISES.items():
        if kw in base_l and not ioc["pais_victima"]:
            ioc["pais_victima"] = pais
            break

    # País origen del ataque (atribución explícita)
    attr_patterns = [
        r'(?:linked to|attributed to|believed to be from|sponsored by|'
        r'atribuido a|vinculado a|apoyado por)\s+(\w[\w\s]{1,20})',
    ]
    for pat in attr_patterns:
        m = re.search(pat, full, re.I)
        if m:
            attr = m.group(1).strip().lower()
            for kw, pais in PAISES.items():
                if kw in attr:
                    ioc["pais_origen"] = pais
                    break
        if ioc["pais_origen"]:
            break

    # Si no hay atribución explícita, inferir del grupo APT conocido
    if not ioc["pais_origen"] and ioc["grupos"]:
        origen_por_grupo = {
            "Lazarus": "Corea del Norte", "Hidden Cobra": "Corea del Norte",
            "Kimsuky": "Corea del Norte", "Velvet Chollima": "Corea del Norte",
            "Fancy Bear": "Rusia", "Apt28": "Rusia", "Cozy Bear": "Rusia",
            "Apt29": "Rusia", "Sandworm": "Rusia", "Turla": "Rusia",
            "Gamaredon": "Rusia", "Forest Blizzard": "Rusia",
            "Midnight Blizzard": "Rusia", "Nobelium": "Rusia",
            "Mustang Panda": "China", "Apt41": "China", "Volt Typhoon": "China",
            "Salt Typhoon": "China", "Bronze President": "China",
            "Charming Kitten": "Irán", "Apt35": "Irán",
            "Muddywater": "Irán", "Mint Sandstorm": "Irán",
        }
        for g in ioc["grupos"]:
            if g in origen_por_grupo:
                ioc["pais_origen"] = origen_por_grupo[g]
                break

    # Impacto resumido (1 línea con los datos más relevantes)
    partes_impacto = []
    if ioc["datos_robados"]:
        partes_impacto.append(ioc["datos_robados"])
    if ioc["sistemas_afectados"]:
        partes_impacto.append(ioc["sistemas_afectados"])
    if ioc["rescate"]:
        partes_impacto.append(f"Rescate: {ioc['rescate']}")
    if partes_impacto:
        ioc["impacto_resumen"] = " · ".join(partes_impacto)[:120]

    return ioc

def detectar_mitre(titulo, resumen):
    texto = (titulo + " " + resumen).lower()
    encontradas = []
    for tac_id, keywords in MITRE_MAP.items():
        for kw in keywords:
            if kw in texto:
                encontradas.append(tac_id)
                break
    if not encontradas:
        return None, None, []
    primario = encontradas[0]
    return primario, MITRE_TACTICS.get(primario, ""), encontradas

def nivel_severidad(cat, mitre_ids, ioc=None):
    criticos = {"TA0040", "TA0006", "TA0043"}
    if cat in ("RANSOMWARE", "INFRAESTRUCTURA CRÍTICA", "APT / ESPIONAJE"):
        return "CRÍTICO"
    if mitre_ids and any(m in criticos for m in mitre_ids):
        return "CRÍTICO"
    if ioc and ioc.get("cvss_score"):
        try:
            if float(ioc["cvss_score"]) >= 9.0:
                return "CRÍTICO"
            if float(ioc["cvss_score"]) >= 7.0:
                return "ALTO"
        except Exception:
            pass
    if cat in ("VULNERABILIDADES", "MALWARE / TROYANOS"):
        return "ALTO"
    if mitre_ids and any(m in {"TA0002", "TA0003", "TA0004", "TA0011"} for m in mitre_ids):
        return "ALTO"
    if cat in ("FILTRACIÓN DE DATOS", "PHISHING / INGENIERÍA SOCIAL"):
        return "MEDIO"
    return "BAJO"

# ─────────────────────────────────────────────────────────────
# CLASIFICACIÓN
# ─────────────────────────────────────────────────────────────
def clasificar_cyber(titulo, resumen):
    texto = (titulo + " " + resumen).lower()
    for ex in EXCLUIR:
        if ex in texto:
            return None, None
    for cat in CATEGORIAS_ORDEN:
        for kw in CATEGORIAS[cat]:
            if kw.lower() in texto:
                return cat, "keywords"
    return None, "revisar"

def _similitud(t1, t2):
    stop = {"de","la","el","los","las","en","a","y","por","con","del",
            "the","of","and","in","to","for","by","on","at","an","is","was"}
    def w(t):
        return {x for x in re.findall(r'[a-záéíóúüña-z0-9]+', t.lower())
                if x not in stop and len(x) > 3}
    p1, p2 = w(t1), w(t2)
    if not p1 or not p2: return 0.0
    return len(p1 & p2) / max(len(p1), len(p2))

def deduplicar(arts):
    unicos = []
    for art in arts:
        dup = False
        for u in unicos:
            if _similitud(art["title"], u["title"]) >= _UMBRAL_DUP:
                fuentes = u.get("_fuentes", [u.get("region", "")])
                if art.get("region", "") not in fuentes:
                    fuentes.append(art.get("region", ""))
                u["_fuentes"] = fuentes
                dup = True
                break
        if not dup:
            art["_fuentes"] = [art.get("region", "")]
            unicos.append(art)
    return unicos

# ─────────────────────────────────────────────────────────────
# ESTADÍSTICAS DE FEEDS
# ─────────────────────────────────────────────────────────────
def stats_cargar():
    if os.path.exists(_stats_path):
        try:
            return json.load(open(_stats_path, encoding="utf-8"))
        except Exception:
            pass
    return {}

def stats_guardar(s):
    with open(_stats_path, "w", encoding="utf-8") as f:
        json.dump(s, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────────────────────
# EXCEL THREAT INTEL
# ─────────────────────────────────────────────────────────────
def acumular_ioc(noticias):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles.differential import DifferentialStyle

    # ── Paleta ──
    _BG_HDR    = "0A1A2E"   # header oscuro
    _FG_HDR    = "00E5FF"   # cyan neon
    _BG_SEV = {
        "CRÍTICO": ("3D0010", "FF4070"),
        "ALTO":    ("2A1400", "FF8030"),
        "MEDIO":   ("1E1800", "FFD030"),
        "BAJO":    ("081A08", "40CC60"),
    }
    _BG_CAT = {
        "RANSOMWARE":                   ("2A0008", "FF6080"),
        "VULNERABILIDADES":             ("1E1000", "FF9040"),
        "APT / ESPIONAJE":              ("150025", "C060FF"),
        "FILTRACIÓN DE DATOS":          ("001520", "40B0FF"),
        "MALWARE / TROYANOS":           ("200010", "FF50A0"),
        "INFRAESTRUCTURA CRÍTICA":      ("220010", "FF4090"),
        "PHISHING / INGENIERÍA SOCIAL": ("1A1000", "FFA020"),
        "CIBERCRIMEN / DARKWEB":        ("080F08", "70FF50"),
    }
    _THIN = Side(style="thin", color="1A2A40")
    _border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _GRAY_ROW = "080F1A"   # fila par

    def _fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def _font(color="C8D8F0", bold=False, size=9):
        return Font(name="Consolas", color=color, bold=bold, size=size)

    def _align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center",
                         wrap_text=wrap, shrink_to_fit=False)

    # ── Abrir o crear workbook ──
    if os.path.exists(_ioc_path):
        wb = openpyxl.load_workbook(_ioc_path)
        ws = wb["Threat Intel"] if "Threat Intel" in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Threat Intel"

        # ── Fila de título principal ──
        ws.merge_cells("A1:AC1")
        title_cell = ws["A1"]
        title_cell.value = "RMSecurity · THREAT INTELLIGENCE — Registro de Amenazas de Ciberseguridad"
        title_cell.font  = Font(name="Consolas", bold=True, size=13, color=_FG_HDR)
        title_cell.fill  = _fill(_BG_HDR)
        title_cell.alignment = _align("center")
        ws.row_dimensions[1].height = 28

        # ── Fila de subtítulo / metadatos ──
        ws.merge_cells("A2:AC2")
        sub = ws["A2"]
        sub.value = f"Generado por: RMSecurity Cyber Intel  ·  Rodrigo Moses  ·  linkedin.com/in/rodrigo-m-793b36152"
        sub.font  = Font(name="Consolas", italic=True, size=8, color="4A6888")
        sub.fill  = _fill("05101E")
        sub.alignment = _align("center")
        ws.row_dimensions[2].height = 16

        # ── Headers de columnas (fila 3) ──
        HEADERS = [
            ("Fecha",             12, "center"),
            ("Severidad",         12, "center"),
            ("Categoría",         22, "left"),
            ("Táctica MITRE",     20, "left"),
            ("ID MITRE",          10, "center"),
            ("Técnicas T1",       16, "left"),
            ("CVSS",               7, "center"),
            ("CVE",               18, "left"),
            ("Actor / Grupo APT", 22, "left"),
            ("País Origen",       14, "center"),
            ("Víctima",           22, "left"),
            ("Sector",            16, "left"),
            ("País Víctima",      14, "center"),
            ("Impacto",           28, "left"),
            ("Datos Robados",     20, "left"),
            ("Sistemas Afect.",   18, "left"),
            ("Rescate / Monto",   18, "left"),
            ("Software",          20, "left"),
            ("Versiones",         12, "left"),
            ("IPs C2",            22, "left"),
            ("Puertos",           12, "left"),
            ("Dominios Mal.",     22, "left"),
            ("Hashes",            26, "left"),
            ("Wallets Crypto",    22, "left"),
            ("Emails",            22, "left"),
            ("Título",            40, "left"),
            ("Resumen",           50, "left"),
            ("Fuente",            18, "left"),
            ("URL",               14, "center"),
        ]
        for col_idx, (hdr, width, align) in enumerate(HEADERS, 1):
            c = ws.cell(row=3, column=col_idx, value=hdr)
            c.font      = Font(name="Consolas", bold=True, size=9, color=_FG_HDR)
            c.fill      = _fill(_BG_HDR)
            c.alignment = _align(align, wrap=True)
            c.border    = _border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[3].height = 26

        ws.freeze_panes  = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"
        ws.sheet_view.showGridLines = False

    # ── Leer existentes (dedup por título) ──
    DATA_ROW = 4  # primera fila de datos
    existentes = set()
    for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
        if row[25]:   # columna "Título" (índice 25)
            existentes.add(str(row[25])[:80])

    nuevas     = 0
    row_actual = ws.max_row + 1 if ws.max_row >= DATA_ROW else DATA_ROW
    is_pair    = (row_actual % 2 == 0)

    for n in noticias:
        key = str(n.get("title", ""))[:80]
        if key in existentes:
            continue

        sev  = n.get("severidad", "BAJO")
        cat  = n.get("cat", "")
        link = n.get("link", "")

        valores = [
            n.get("fecha_pub", str(date.today())),
            sev,
            cat,
            n.get("mitre_name", ""),
            n.get("mitre_id", ""),
            n.get("tecnicas", ""),
            n.get("cvss", ""),
            n.get("cve", ""),
            n.get("actor", ""),
            n.get("pais_origen", ""),
            n.get("victima", ""),
            n.get("sector", ""),
            n.get("pais_victima", ""),
            n.get("impacto", ""),
            n.get("datos_robados", ""),
            n.get("sistemas_afect", ""),
            n.get("rescate", ""),
            n.get("software", ""),
            n.get("versiones", ""),
            n.get("ips", ""),
            n.get("puertos", ""),
            n.get("dominios", ""),
            n.get("hashes", ""),
            n.get("wallets", ""),
            n.get("emails", ""),
            n.get("title", ""),
            n.get("resumen", "")[:300],
            n.get("region", ""),
            "→ Abrir" if link else "",
        ]

        # Fondo alternado por fila
        bg_fila = _GRAY_ROW if is_pair else "040C18"

        for col_idx, val in enumerate(valores, 1):
            c = ws.cell(row=row_actual, column=col_idx, value=val)
            c.font      = _font()
            c.border    = _border
            c.alignment = _align("center" if col_idx in (1,2,5,7,10,13,29) else "left",
                                  wrap=(col_idx in (14,15,16,17,26,27)))

            # Fondo base alternado
            c.fill = _fill(bg_fila)

        # Colorear celda SEVERIDAD (col 2)
        sev_bg, sev_fg = _BG_SEV.get(sev, ("080F1A", "888888"))
        c_sev = ws.cell(row=row_actual, column=2)
        c_sev.fill = _fill(sev_bg)
        c_sev.font = Font(name="Consolas", bold=True, size=9, color=sev_fg)

        # Colorear celda CATEGORÍA (col 3)
        cat_bg, cat_fg = _BG_CAT.get(cat, ("080F1A", "88AACC"))
        c_cat = ws.cell(row=row_actual, column=3)
        c_cat.fill = _fill(cat_bg)
        c_cat.font = Font(name="Consolas", size=9, color=cat_fg)

        # Colorear celda PAÍS ORIGEN (col 10) — rojo si es país hostil
        pais_o = n.get("pais_origen", "")
        if pais_o in ("Rusia", "China", "Irán", "Corea del Norte"):
            ws.cell(row=row_actual, column=10).fill = _fill("2A0008")
            ws.cell(row=row_actual, column=10).font = Font(
                name="Consolas", size=9, color="FF4060", bold=True)

        # CVE en naranja si existe
        if n.get("cve"):
            ws.cell(row=row_actual, column=8).font = Font(
                name="Consolas", size=9, color="FF8030")

        # CVSS en rojo si ≥ 9, naranja si ≥ 7
        try:
            cvss_val = float(n.get("cvss") or 0)
            c_cvss = ws.cell(row=row_actual, column=7)
            if cvss_val >= 9.0:
                c_cvss.font = Font(name="Consolas", bold=True, size=9, color="FF2050")
                c_cvss.fill = _fill("2A0010")
            elif cvss_val >= 7.0:
                c_cvss.font = Font(name="Consolas", bold=True, size=9, color="FF7020")
        except (ValueError, TypeError):
            pass

        # URL como hipervínculo (col 29)
        if link:
            c_url = ws.cell(row=row_actual, column=29)
            c_url.hyperlink = link
            c_url.font = Font(name="Consolas", size=9, color="00AAFF",
                              underline="single")

        # Actor en rojo si es APT conocido
        if n.get("actor"):
            ws.cell(row=row_actual, column=9).font = Font(
                name="Consolas", size=9, color="FF6080")

        ws.row_dimensions[row_actual].height = 18
        existentes.add(key)
        nuevas   += 1
        is_pair   = not is_pair
        row_actual += 1

    # ── Hoja de resumen ──
    _escribir_resumen(wb, ws)

    wb.save(_ioc_path)
    return nuevas


def _escribir_resumen(wb, ws_data):
    """Genera o actualiza la hoja 'Resumen' con estadísticas del registro."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    nombre = "Resumen"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre, 0)   # primera hoja
    ws.sheet_view.showGridLines = False

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(color="C8D8F0", bold=False, size=9):
        return Font(name="Consolas", color=color, bold=bold, size=size)

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"].value     = "RMSecurity · Threat Intelligence — Resumen Ejecutivo"
    ws["A1"].font      = Font(name="Consolas", bold=True, size=14, color="00E5FF")
    ws["A1"].fill      = _fill("040C18")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Actualizado: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font      = Font(name="Consolas", italic=True, size=8, color="4A6888")
    ws["A2"].fill      = _fill("030810")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Leer datos de la hoja principal (fila 4 en adelante)
    DATA_ROW = 4
    conteo_sev = {"CRÍTICO": 0, "ALTO": 0, "MEDIO": 0, "BAJO": 0}
    conteo_cat = {}
    conteo_pais_o = {}
    conteo_actor  = {}
    total = 0

    for row in ws_data.iter_rows(min_row=DATA_ROW, values_only=True):
        if not row[25]:   # sin título = fila vacía
            continue
        total += 1
        sev = str(row[1] or "")
        cat = str(row[2] or "")
        po  = str(row[9] or "")
        act = str(row[8] or "")
        if sev in conteo_sev:
            conteo_sev[sev] += 1
        if cat:
            conteo_cat[cat] = conteo_cat.get(cat, 0) + 1
        if po:
            conteo_pais_o[po] = conteo_pais_o.get(po, 0) + 1
        if act:
            for a in act.split(","):
                a = a.strip()
                if a:
                    conteo_actor[a] = conteo_actor.get(a, 0) + 1

    # KPI cards — fila 4
    _SEV_COLORS = {
        "CRÍTICO": ("3D0010", "FF4070"),
        "ALTO":    ("2A1400", "FF8030"),
        "MEDIO":   ("1E1800", "FFD030"),
        "BAJO":    ("081A08", "40CC60"),
    }
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 14

    kpi_cols = [("A", "TOTAL", str(total), "0A1A2E", "00E5FF"),
                ("B", "CRÍTICO", str(conteo_sev["CRÍTICO"]), "3D0010", "FF4070"),
                ("C", "ALTO",    str(conteo_sev["ALTO"]),    "2A1400", "FF8030"),
                ("D", "MEDIO",   str(conteo_sev["MEDIO"]),   "1E1800", "FFD030"),
                ("E", "BAJO",    str(conteo_sev["BAJO"]),    "081A08", "40CC60")]

    for col_l, label, valor, bg, fg in kpi_cols:
        ws[f"{col_l}4"].fill = _fill(bg)
        ws[f"{col_l}4"].value = ""
        ws[f"{col_l}5"].value = valor
        ws[f"{col_l}5"].font = Font(name="Consolas", bold=True, size=22, color=fg)
        ws[f"{col_l}5"].fill = _fill(bg)
        ws[f"{col_l}5"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_l}6"].value = label
        ws[f"{col_l}6"].font  = Font(name="Consolas", size=8, color=fg)
        ws[f"{col_l}6"].fill  = _fill(bg)
        ws[f"{col_l}6"].alignment = Alignment(horizontal="center")
        ws[f"{col_l}7"].fill = _fill(bg)
        ws[f"{col_l}7"].value = ""
        ws.column_dimensions[col_l].width = 16

    # Tabla por Categoría — columna A, fila 9
    ws["A9"].value = "Por Categoría"
    ws["A9"].font  = Font(name="Consolas", bold=True, size=10, color="00E5FF")
    ws["A9"].fill  = _fill("040C18")
    ws["B9"].value = "Cantidad"
    ws["B9"].font  = _font("4A6888", bold=True)
    ws["B9"].fill  = _fill("040C18")

    _CAT_FG = {
        "RANSOMWARE": "FF6080", "VULNERABILIDADES": "FF9040",
        "APT / ESPIONAJE": "C060FF", "FILTRACIÓN DE DATOS": "40B0FF",
        "MALWARE / TROYANOS": "FF50A0", "INFRAESTRUCTURA CRÍTICA": "FF4090",
        "PHISHING / INGENIERÍA SOCIAL": "FFA020", "CIBERCRIMEN / DARKWEB": "70FF50",
    }
    sorted_cat = sorted(conteo_cat.items(), key=lambda x: -x[1])
    for i, (cat, cnt) in enumerate(sorted_cat, 10):
        bg_c = "080F1A" if i % 2 == 0 else "040C14"
        ws[f"A{i}"].value = cat
        ws[f"A{i}"].font  = Font(name="Consolas", size=9,
                                  color=_CAT_FG.get(cat, "AABBCC"))
        ws[f"A{i}"].fill  = _fill(bg_c)
        ws[f"A{i}"].alignment = Alignment(horizontal="left")
        ws[f"B{i}"].value = cnt
        ws[f"B{i}"].font  = Font(name="Consolas", bold=True, size=9,
                                  color=_CAT_FG.get(cat, "AABBCC"))
        ws[f"B{i}"].fill  = _fill(bg_c)
        ws[f"B{i}"].alignment = Alignment(horizontal="center")

    # Tabla por País de origen — columna D, fila 9
    ws["D9"].value = "País de Origen"
    ws["D9"].font  = Font(name="Consolas", bold=True, size=10, color="00E5FF")
    ws["D9"].fill  = _fill("040C18")
    ws["E9"].value = "Amenazas"
    ws["E9"].font  = _font("4A6888", bold=True)
    ws["E9"].fill  = _fill("040C18")

    _PAIS_FG = {"Rusia": "FF4060", "China": "FF6020",
                "Irán": "FF9020", "Corea del Norte": "C040FF"}
    sorted_pais = sorted(conteo_pais_o.items(), key=lambda x: -x[1])
    for i, (pais, cnt) in enumerate(sorted_pais[:10], 10):
        bg_c = "080F1A" if i % 2 == 0 else "040C14"
        ws[f"D{i}"].value = pais
        ws[f"D{i}"].font  = Font(name="Consolas", size=9,
                                  color=_PAIS_FG.get(pais, "AABBCC"))
        ws[f"D{i}"].fill  = _fill(bg_c)
        ws["D" + str(i)].alignment = Alignment(horizontal="left")
        ws[f"E{i}"].value = cnt
        ws[f"E{i}"].font  = Font(name="Consolas", bold=True, size=9,
                                  color=_PAIS_FG.get(pais, "AABBCC"))
        ws[f"E{i}"].fill  = _fill(bg_c)
        ws[f"E{i}"].alignment = Alignment(horizontal="center")
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12

    # Tabla actores más frecuentes — columna G, fila 9
    ws["G9"].value = "Actor / Grupo APT"
    ws["G9"].font  = Font(name="Consolas", bold=True, size=10, color="00E5FF")
    ws["G9"].fill  = _fill("040C18")
    ws["H9"].value = "Menciones"
    ws["H9"].font  = _font("4A6888", bold=True)
    ws["H9"].fill  = _fill("040C18")

    sorted_act = sorted(conteo_actor.items(), key=lambda x: -x[1])[:12]
    for i, (act, cnt) in enumerate(sorted_act, 10):
        bg_c = "080F1A" if i % 2 == 0 else "040C14"
        ws[f"G{i}"].value = act
        ws[f"G{i}"].font  = Font(name="Consolas", size=9, color="FF6080")
        ws[f"G{i}"].fill  = _fill(bg_c)
        ws[f"G{i}"].alignment = Alignment(horizontal="left")
        ws[f"H{i}"].value = cnt
        ws[f"H{i}"].font  = Font(name="Consolas", bold=True, size=9, color="FF6080")
        ws[f"H{i}"].fill  = _fill(bg_c)
        ws[f"H{i}"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["F"].width = 2  # separador

    # Fondo negro en toda la hoja (celdas no usadas)
    for r in range(1, 30):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = PatternFill("solid", fgColor="02060F")

    # Mover Threat Intel al final
    wb.move_sheet("Threat Intel", offset=len(wb.sheetnames))

# ─────────────────────────────────────────────────────────────
# PARSEO DE FECHA
# ─────────────────────────────────────────────────────────────
def parsear_fecha(pub_str):
    if not pub_str:
        return date.today()
    for fmt in ("%Y-%m-%d", "%a, %d %b %Y", "%d %b %Y"):
        try:
            return datetime.strptime(pub_str[:20].strip(), fmt).date()
        except Exception:
            pass
    return date.today()

# ─────────────────────────────────────────────────────────────
# APP PRINCIPAL
# ─────────────────────────────────────────────────────────────
class CyberIntelApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("RMSecurity · Threat Intelligence de Ciberseguridad")
        self.geometry("1400x860")
        self.minsize(1050, 650)
        self.configure(bg=C["bg0"])
        self.resizable(True, True)
        try:
            self.iconbitmap(_resource("dacc_eagle.ico"))
        except Exception:
            pass

        self.noticias_data = {}
        self.noticias_vars = {}
        self._id_counter   = itertools.count()
        self._buscando     = False
        self._usar_ollama  = ollama_disponible()
        self._sort_col     = "fecha"
        self._sort_desc    = True  # más nuevo primero por defecto

        self._build_ui()
        self.center()
        self.after(400, self._check_ollama_status)

    def center(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - self.winfo_width())  // 2
        y = (self.winfo_screenheight() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _check_ollama_status(self):
        ok = ollama_disponible()
        self._usar_ollama = ok
        if ok:
            self.lbl_ollama.config(text="⬡ Ollama: ON", fg=C["green"])
        else:
            self.lbl_ollama.config(text="⬡ Ollama: OFF", fg=C["amber"])

    # ─── UI ───
    def _build_ui(self):
        self._estilo_dark()

        # TOPBAR
        top = tk.Frame(self, bg="#010408", pady=7)
        top.pack(fill="x")
        self._lbl_title_top = tk.Label(top, text=T("title_top"),
                 bg="#010408", fg=C["cyan"],
                 font=("Consolas", 14, "bold"))
        self._lbl_title_top.pack(side="left", padx=14)

        right_top = tk.Frame(top, bg="#010408")
        right_top.pack(side="right", padx=8)
        self.lbl_ollama = tk.Label(right_top, text="⬡ Ollama: ...",
                                   bg="#010408", fg=C["amber"],
                                   font=("Consolas", 9))
        self.lbl_ollama.pack(side="right", padx=10)
        self._btn_excel_top = self._btn(right_top, T("btn_excel"), C["blue"],  self._abrir_ioc,  lado="right")
        self._btn_feeds_top = self._btn(right_top, T("btn_feeds"), C["border"], self._ver_feeds, lado="right", fg=C["txt"])
        self._btn_lang = tk.Button(right_top, text=T("lang_btn"),
                                   bg=C["bg2"], fg=C["txt_hi"],
                                   font=("Consolas", 9, "bold"),
                                   relief="flat", padx=8, pady=3,
                                   activebackground=C["bg3"], activeforeground=C["cyan"],
                                   command=self._toggle_lang)
        self._btn_lang.pack(side="right", padx=4)

        # SEPARADOR NEON
        sep = tk.Frame(self, bg=C["cyan"], height=1)
        sep.pack(fill="x")

        # BARRA DE FILTROS
        filtros = tk.Frame(self, bg=C["bg1"], pady=5)
        filtros.pack(fill="x", padx=6, pady=(2, 0))

        self._lbl_cat_filter = self._lbl_ret("Categoría:", filtros, lado="left")
        self._cat_filter = tk.StringVar(value=T("all"))
        self._combo_cat = ttk.Combobox(filtros, textvariable=self._cat_filter,
                              values=[T("all")] + CATEGORIAS_ORDEN,
                              state="readonly", width=26, font=("Consolas", 9))
        self._combo_cat.pack(side="left", padx=4)
        self._combo_cat.bind("<<ComboboxSelected>>", lambda e: self._aplicar_filtro())

        self._lbl_sev_filter = self._lbl_ret("Severidad:", filtros, lado="left", pad=(10, 4))
        self._sev_filter = tk.StringVar(value=T("all"))
        self._combo_sev = ttk.Combobox(filtros, textvariable=self._sev_filter,
                                  values=T("sev_values"),
                                  state="readonly", width=9, font=("Consolas", 9))
        self._combo_sev.pack(side="left", padx=4)
        self._combo_sev.bind("<<ComboboxSelected>>", lambda e: self._aplicar_filtro())

        self._lbl_buscar = self._lbl_ret("Buscar:", filtros, lado="left", pad=(10, 4))
        self._txt_search = tk.Entry(filtros, bg=C["bg2"], fg=C["txt_hi"],
                                    insertbackground=C["cyan"],
                                    font=("Consolas", 10), width=24,
                                    relief="flat", bd=0,
                                    highlightthickness=1,
                                    highlightcolor=C["cyan"],
                                    highlightbackground=C["border"])
        self._txt_search.pack(side="left", padx=4, ipady=3)
        self._txt_search.bind("<Return>", lambda e: self._aplicar_filtro())

        tk.Button(filtros, text="⌕", bg=C["bg2"], fg=C["cyan"],
                  font=("Consolas", 10, "bold"), relief="flat", padx=6,
                  activebackground=C["bg3"], activeforeground=C["cyan"],
                  command=self._aplicar_filtro).pack(side="left", padx=2)

        self._lbl_count = tk.Label(filtros, text="", bg=C["bg1"],
                                    fg=C["txt"], font=("Consolas", 9))
        self._lbl_count.pack(side="right", padx=12)

        # BARRA DE ACCIONES
        acc = tk.Frame(self, bg=C["bg0"], pady=4)
        acc.pack(fill="x", padx=6)

        self._btn_buscar  = self._btn(acc, T("btn_search"),   C["cyan"],   self._iniciar_busqueda, bold=True, fg=C["bg0"])
        self._btn_guardar = self._btn(acc, T("btn_save_ioc"), "#1A003A",   self._guardar_ioc,      fg=C["violet"], bold=True)
        self._btn_limpiar = self._btn(acc, T("btn_clear"),    C["bg1"],    self._limpiar,          fg=C["txt"])
        self._btn_info    = self._btn(acc, "ℹ Info",          C["bg2"],    self._mostrar_info,     fg=C["cyan"], lado="right")

        self._tooltip(self._btn_buscar,  "Buscar noticias en los 47 feeds RSS")
        self._tooltip(self._btn_guardar, "Exportar noticias seleccionadas a Excel (IoC)")
        self._tooltip(self._btn_limpiar, "Limpiar resultados de la lista")
        self._tooltip(self._btn_info,    "Cómo usar el programa")
        self._tooltip(self._btn_excel_top, "Abrir reporte Excel existente")
        self._tooltip(self._btn_feeds_top, "Ver lista de feeds RSS monitoreados")
        self._tooltip(self._btn_lang,      "Cambiar idioma / Switch language")

        # PANEL PRINCIPAL
        main = tk.PanedWindow(self, orient="horizontal", bg=C["bg0"],
                               sashwidth=3, sashrelief="flat",
                               sashpad=0)
        main.pack(fill="both", expand=True, padx=4, pady=4)
        self._main_paned = main

        # ── LISTA IZQUIERDA ──
        izq = tk.Frame(main, bg=C["bg0"])
        main.add(izq, minsize=700)

        # Columnas: fecha primero
        cols = ("fecha", "sel", "sev", "cat", "mitre", "titulo", "actor", "fuente")
        self.tree = ttk.Treeview(izq, columns=cols, show="headings",
                                  selectmode="extended")

        hdrs = {
            "fecha":  (T("col_date"),   95,  False),
            "sev":    (T("col_sev"),    78,  False),
            "cat":    (T("col_cat"),    155, False),
            "mitre":  (T("col_mitre"), 165, False),
            "titulo": (T("col_title"), 330, True),
            "actor":  (T("col_actor"), 125, False),
            "fuente": (T("col_source"),110, False),
        }
        self._tree_hdrs = hdrs
        for col, (txt, w, stretch) in hdrs.items():
            self.tree.heading(col, text=txt,
                              command=lambda c=col: self._ordenar(c))
            self.tree.column(col, width=w, stretch=stretch,
                             anchor="center" if col in ("sev","fecha") else "w")

        # Tags neon por categoría (fondo oscuro, texto neon)
        for cat, (bg, fg) in CAT_COLORS.items():
            safe = cat.replace(" ", "_").replace("/", "_")
            self.tree.tag_configure(safe, background=bg, foreground=fg)

        # Tags por severidad — solo para la columna sev, usamos el mismo tag
        self.tree.tag_configure("SEV_CRÍTICO", foreground=C["red"],    background="#120005")
        self.tree.tag_configure("SEV_ALTO",    foreground=C["orange"], background="#100800")
        self.tree.tag_configure("SEV_MEDIO",   foreground=C["yellow"], background="#0E0C00")
        self.tree.tag_configure("SEV_BAJO",    foreground=C["green"],  background="#050F05")
        self.tree.tag_configure("seleccionada", background="#0A2A10",   foreground=C["green"])
        self.tree.tag_configure("match_pais",   background="#1A1500",   foreground=C["yellow"])

        sb_y = ttk.Scrollbar(izq, orient="vertical",   command=self.tree.yview)
        sb_x = ttk.Scrollbar(izq, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        izq.grid_rowconfigure(0, weight=1)
        izq.grid_columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-Button-1>",  self._abrir_link_double)

        # ── PANEL DERECHO - DETALLE ──
        der = tk.Frame(main, bg=C["bg1"])
        main.add(der, minsize=310)

        # Header detalle
        tk.Frame(der, bg=C["cyan"], height=1).pack(fill="x")
        hdr_row = tk.Frame(der, bg=C["bg1"])
        hdr_row.pack(fill="x", padx=8, pady=(4, 2))
        self._lbl_analysis_hdr = tk.Label(hdr_row, text=T("panel_analysis"), bg=C["bg1"], fg=C["cyan"],
                 font=("Consolas", 9, "bold"), anchor="w")
        self._lbl_analysis_hdr.pack(side="left")
        self._btn_informe = tk.Button(hdr_row, text="📋 Informe",
                  bg="#0A1020", fg=C["amber"],
                  font=("Consolas", 8, "bold"), relief="flat", padx=6, pady=1,
                  activebackground=C["bg3"], activeforeground=C["yellow"],
                  command=self._copiar_informe_rapido)
        self._btn_informe.pack(side="right")
        self._tooltip(self._btn_informe, "Copiar informe completo al portapapeles")

        # Badges sev + cat
        self._lbl_sev = tk.Label(der, text="", bg=C["bg1"], fg="white",
                                  font=("Consolas", 11, "bold"), pady=3)
        self._lbl_sev.pack(fill="x", padx=8)

        self._lbl_cat = tk.Label(der, text="", bg=C["bg1"], fg=C["cyan"],
                                  font=("Consolas", 9, "bold"), wraplength=295,
                                  anchor="w", pady=2)
        self._lbl_cat.pack(fill="x", padx=8)

        # MITRE block
        mitre_frame = tk.Frame(der, bg="#08001A", highlightbackground=C["violet"],
                                highlightthickness=1)
        mitre_frame.pack(fill="x", padx=8, pady=4)
        tk.Label(mitre_frame, text="MITRE ATT&CK", bg="#08001A", fg=C["violet"],
                 font=("Consolas", 7, "bold")).pack(anchor="w", padx=6, pady=(3, 0))
        self._lbl_mitre = tk.Label(mitre_frame, text="—", bg="#08001A", fg="#D070FF",
                                    font=("Consolas", 9), wraplength=285,
                                    justify="left", anchor="w")
        self._lbl_mitre.pack(fill="x", padx=6, pady=(0, 5))

        # IoC block
        ioc_frame = tk.Frame(der, bg="#001008", highlightbackground=C["green"],
                              highlightthickness=1)
        ioc_frame.pack(fill="x", padx=8, pady=(0, 4))
        self._lbl_ioc_hdr = tk.Label(ioc_frame, text=T("panel_ioc"), bg="#001008", fg=C["green"],
                 font=("Consolas", 7, "bold"))
        self._lbl_ioc_hdr.pack(anchor="w", padx=6, pady=(3, 0))

        self._ioc_labels = {}
        self._ioc_field_keys = [
            ("actor",    "ioc_actor",    C["red"]),
            ("pais_o",   "ioc_origin",   C["red"]),
            ("victima",  "ioc_victim",   C["amber"]),
            ("sector",   "ioc_sector",   C["amber"]),
            ("pais_v",   "ioc_country_v",C["txt_hi"]),
            ("impacto",  "ioc_impact",   C["orange"]),
            ("datos",    "ioc_data",     C["orange"]),
            ("sistemas", "ioc_systems",  C["orange"]),
            ("rescate",  "ioc_ransom",   C["red"]),
            ("cve",      "CVE",          C["orange"]),
            ("cvss",     "CVSS",         C["red"]),
            ("software", "ioc_software", C["cyan"]),
            ("versiones","ioc_versions", C["cyan"]),
            ("tecnicas", "ioc_techniques",C["violet"]),
            ("ips",      "ioc_ips",      C["cyan"]),
            ("puertos",  "ioc_ports",    C["txt_hi"]),
            ("dominios", "ioc_domains",  C["orange"]),
            ("hashes",   "ioc_hashes",   "#888888"),
            ("wallets",  "ioc_wallets",  C["amber"]),
            ("emails",   "ioc_emails",   C["txt"]),
        ]
        ioc_fields = [(k, T(tkey) if tkey.startswith("ioc_") else tkey, col)
                      for k, tkey, col in self._ioc_field_keys]
        # Contenedor scrollable para el bloque IoC
        ioc_inner = tk.Frame(ioc_frame, bg="#001008")
        ioc_inner.pack(fill="x", padx=0)
        for key, label, color in ioc_fields:
            row = tk.Frame(ioc_inner, bg="#001008")
            row.pack(fill="x", padx=6, pady=1)
            tk.Label(row, text=f"{label}:", bg="#001008", fg="#556677",
                     font=("Consolas", 7), width=13, anchor="e").pack(side="left")
            lbl = tk.Label(row, text="—", bg="#001008", fg=color,
                           font=("Consolas", 8), anchor="w", wraplength=170,
                           justify="left", cursor="hand2")
            lbl.pack(side="left", padx=3)
            lbl.bind("<Button-1>", lambda e, w=lbl, k=key: self._copiar_ioc_field(w, k))
            self._tooltip(lbl, "Clic para copiar")
            self._ioc_labels[key] = lbl

        # Resumen
        tk.Frame(der, bg=C["border"], height=1).pack(fill="x", padx=8, pady=(4, 0))
        self._lbl_summary_hdr = tk.Label(der, text=T("panel_summary"), bg=C["bg1"], fg=C["txt"],
                 font=("Consolas", 7, "bold"), anchor="w")
        self._lbl_summary_hdr.pack(fill="x", padx=8, pady=(3, 0))
        self._txt_detail = scrolledtext.ScrolledText(
            der, bg="#050D18", fg=C["txt_hi"],
            font=("Consolas", 9), wrap="word",
            relief="flat", bd=0,
            insertbackground=C["cyan"],
            highlightthickness=0)
        self._txt_detail.pack(fill="both", expand=True, padx=8, pady=(2, 4))

        self._btn_open_link = tk.Button(der, text=T("btn_open_link"),
                  bg=C["bg2"], fg=C["cyan"], font=("Consolas", 9),
                  relief="flat", padx=8, pady=3,
                  activebackground=C["bg3"], activeforeground=C["cyan"],
                  command=self._abrir_link).pack(pady=(0, 6))

        # ── PANEL DE CARGA (oculto por defecto, aparece durante búsqueda) ──
        self._frame_carga = tk.Frame(self, bg="#020810", pady=12)
        # No se hace pack aquí — se muestra/oculta dinámicamente

        # Barra de progreso ancha y visible
        self._progress = ttk.Progressbar(
            self._frame_carga, length=700, mode="indeterminate",
            style="Neon.Horizontal.TProgressbar")
        self._progress.pack(pady=(0, 8))

        # Texto de estado dentro del panel de carga — grande y neon
        self._lbl_carga_titulo = tk.Label(
            self._frame_carga,
            text=T("loading_title") + "...",
            bg="#020810", fg=C["cyan"],
            font=("Consolas", 13, "bold"))
        self._lbl_carga_titulo.pack()

        self.lbl_status = tk.Label(
            self._frame_carga,
            text=T("loading_feeds"),
            bg="#020810", fg=C["amber"],
            font=("Consolas", 10), anchor="center")
        self.lbl_status.pack(pady=(4, 0))

        self._lbl_carga_sub = tk.Label(
            self._frame_carga,
            text="",
            bg="#020810", fg="#445566",
            font=("Consolas", 8))
        self._lbl_carga_sub.pack(pady=(2, 4))

        # ── BARRA STATUS (siempre visible en el footer) ──
        sep2 = tk.Frame(self, bg=C["border"], height=1)
        sep2.pack(fill="x")
        status_bar = tk.Frame(self, bg="#010408", pady=3)
        status_bar.pack(fill="x")

        self._lbl_status_footer = tk.Label(
            status_bar,
            text=T("status_ready"),
            bg="#010408", fg=C["txt"],
            font=("Consolas", 9), anchor="w")
        self._lbl_status_footer.pack(side="left", padx=10)

        # Contador derecha de la status bar
        self._lbl_status_right = tk.Label(
            status_bar, text="",
            bg="#010408", fg=C["txt"],
            font=("Consolas", 9), anchor="e")
        self._lbl_status_right.pack(side="right", padx=16)

        # Crédito discreto
        lbl_cred = tk.Label(status_bar,
                             text="by Rodrigo Moses · linkedin.com/in/rodrigo-m-793b36152",
                             bg="#010408", fg="#334455",
                             font=("Consolas", 8), cursor="hand2")
        lbl_cred.pack(side="right", padx=16)
        lbl_cred.bind("<Button-1>",
                       lambda e: __import__("webbrowser").open(
                           "https://www.linkedin.com/in/rodrigo-m-793b36152/"))

    def _btn(self, parent, text, bg, cmd, bold=False, fg="white", lado="left"):
        font = ("Consolas", 9, "bold") if bold else ("Consolas", 9)
        b = tk.Button(parent, text=text, bg=bg, fg=fg, font=font,
                      relief="flat", padx=10, pady=4,
                      activebackground=C["bg3"], activeforeground=C["cyan"],
                      command=cmd)
        b.pack(side=lado, padx=4)
        return b

    def _tooltip(self, widget, text):
        tip = None
        def show(e):
            nonlocal tip
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{e.x_root+12}+{e.y_root+18}")
            tk.Label(tip, text=text, bg="#0F1830", fg=C["cyan"],
                     font=("Consolas", 8), padx=6, pady=3,
                     relief="flat", bd=1,
                     highlightbackground=C["border"],
                     highlightthickness=1).pack()
        def hide(e):
            nonlocal tip
            if tip:
                tip.destroy()
                tip = None
        widget.bind("<Enter>", show)
        widget.bind("<Leave>", hide)

    def _lbl(self, text, parent, lado="left", pad=(6, 4)):
        tk.Label(parent, text=text, bg=C["bg1"], fg=C["txt"],
                 font=("Consolas", 9)).pack(side=lado, padx=pad)

    def _lbl_ret(self, text, parent, lado="left", pad=(6, 4)):
        w = tk.Label(parent, text=text, bg=C["bg1"], fg=C["txt"],
                     font=("Consolas", 9))
        w.pack(side=lado, padx=pad)
        return w

    def _toggle_lang(self):
        global _LANG
        _LANG = "EN" if _LANG == "ES" else "ES"
        self._aplicar_idioma()

    def _aplicar_idioma(self):
        self._lbl_title_top.config(text=T("title_top"))
        self.title(T("title_top").replace("◈ ", "RMSecurity · "))
        self._btn_lang.config(text=T("lang_btn"))
        self._btn_buscar.config(text=T("btn_search"))
        self._btn_guardar.config(text=T("btn_save_ioc"))
        self._btn_limpiar.config(text=T("btn_clear"))
        self._btn_excel_top.config(text=T("btn_excel"))
        self._btn_feeds_top.config(text=T("btn_feeds"))
        self._btn_open_link.config(text=T("btn_open_link"))
        self._lbl_cat_filter.config(text=T("lbl_category"))
        self._lbl_sev_filter.config(text=T("lbl_severity"))
        self._lbl_buscar.config(text=T("lbl_search"))
        self._lbl_analysis_hdr.config(text=T("panel_analysis"))
        self._lbl_ioc_hdr.config(text=T("panel_ioc"))
        self._lbl_summary_hdr.config(text=T("panel_summary"))
        self._lbl_status_footer.config(text=T("status_ready"))
        self._lbl_carga_titulo.config(text=T("loading_title") + "...")
        self.lbl_status.config(text=T("loading_feeds"))
        # Columnas del treeview
        col_map = {
            "fecha": "col_date", "sev": "col_sev",
            "cat": "col_cat", "mitre": "col_mitre", "titulo": "col_title",
            "actor": "col_actor", "fuente": "col_source",
        }
        for col, tkey in col_map.items():
            self.tree.heading(col, text=T(tkey))
        # Combo filtros
        cur_cat = self._cat_filter.get()
        cur_sev = self._sev_filter.get()
        self._combo_cat.config(values=[T("all")] + CATEGORIAS_ORDEN)
        self._combo_sev.config(values=T("sev_values"))
        self._cat_filter.set(T("all") if cur_cat in ("TODAS", "ALL") else cur_cat)
        self._sev_filter.set(T("all") if cur_sev in ("TODAS", "ALL") else cur_sev)
        # Labels IoC del panel derecho
        for key, tkey, _ in self._ioc_field_keys:
            if key in self._ioc_labels:
                label_widget = self._ioc_labels[key]
                # Actualizar el label de la etiqueta (el tk.Label hermano izquierdo)
                parent_row = label_widget.master
                children = parent_row.winfo_children()
                if children:
                    lbl_text = T(tkey) if tkey.startswith("ioc_") else tkey
                    children[0].config(text=f"{lbl_text}:")

    def _estilo_dark(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview",
                         background=C["bg2"], foreground=C["txt_hi"],
                         fieldbackground=C["bg2"], rowheight=24,
                         font=("Consolas", 9))
        style.configure("Treeview.Heading",
                         background=C["bg0"], foreground=C["cyan"],
                         font=("Consolas", 9, "bold"),
                         relief="flat")
        style.map("Treeview",
                  background=[("selected", C["sel"])],
                  foreground=[("selected", C["cyan"])])
        style.configure("Neon.Horizontal.TProgressbar",
                         troughcolor=C["bg2"], background=C["cyan"],
                         darkcolor=C["cyan"], lightcolor=C["cyan"])
        style.configure("TCombobox",
                         fieldbackground=C["bg2"], background=C["bg2"],
                         foreground=C["txt_hi"], selectbackground=C["sel"],
                         arrowcolor=C["cyan"])
        style.configure("TScrollbar",
                         troughcolor=C["bg0"], background=C["border"],
                         arrowcolor=C["txt"])

    # ─── ORDENAMIENTO POR COLUMNA ───
    def _ordenar(self, col):
        if self._sort_col == col:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col = col
            self._sort_desc = (col == "fecha")  # fecha: desc por defecto

        items = [(self.tree.set(iid, col), iid) for iid in self.tree.get_children()]

        if col == "fecha":
            items.sort(key=lambda x: x[0], reverse=self._sort_desc)
        else:
            items.sort(key=lambda x: x[0].lower(), reverse=self._sort_desc)

        for idx, (_, iid) in enumerate(items):
            self.tree.move(iid, "", idx)

        arrow = " ↓" if self._sort_desc else " ↑"
        for c in ("fecha", "sev", "cat", "mitre", "titulo", "actor", "fuente"):
            lbl = {
                "fecha": "Fecha", "sev": "Severidad",
                "cat": "Categoría", "mitre": "MITRE ATT&CK",
                "titulo": "Título", "actor": "Actor / Grupo", "fuente": "Fuente",
            }[c]
            self.tree.heading(c, text=lbl + (arrow if c == col else ""))

    # ─── PANEL DE CARGA ───
    def _mostrar_carga(self, msg="Consultando feeds de ciberseguridad..."):
        """Muestra el panel de carga prominente sobre la lista."""
        self._frame_carga.pack(fill="x", padx=0, pady=0,
                                before=self._main_paned)
        self.lbl_status.config(text=msg)
        self._lbl_carga_sub.config(text="")
        self._progress.start(8)
        self._animar_puntos(0)

    def _ocultar_carga(self):
        """Oculta el panel de carga."""
        self._progress.stop()
        self._frame_carga.pack_forget()
        if hasattr(self, "_after_anim"):
            self.after_cancel(self._after_anim)

    def _animar_puntos(self, tick):
        """Anima el título con puntos para dar sensación de actividad."""
        if not self._buscando:
            return
        puntos = "." * (tick % 4)
        self._lbl_carga_titulo.config(text=f"⟳  BUSCANDO AMENAZAS{puntos}")
        self._after_anim = self.after(400, self._animar_puntos, tick + 1)

    def _actualizar_status_carga(self, msg, sub=""):
        """Actualiza el texto dentro del panel de carga (thread-safe vía after)."""
        self.lbl_status.config(text=msg)
        self._lbl_carga_sub.config(text=sub)

    # ─── BÚSQUEDA ───
    def _iniciar_busqueda(self):
        if self._buscando:
            return
        self._buscando = True
        self.noticias_data.clear()
        self.noticias_vars.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._lbl_count.config(text="")
        self._lbl_status_footer.config(text="Buscando...")
        self._mostrar_carga()
        threading.Thread(target=self._buscar_thread, daemon=True).start()

    def _buscar_thread(self):
        stats   = stats_cargar()
        todos   = []
        total_feeds = len(FEEDS)
        feeds_ok    = [0]   # mutable para closure

        def _upd(msg, sub=""):
            self.after(0, lambda m=msg, s=sub: self._actualizar_status_carga(m, s))

        def fetch(nombre, url):
            try:
                feed  = _fetch_feed(url)
                items = []
                for e in feed.entries[:20]:
                    title = limpiar_texto(getattr(e, "title", ""))
                    summ  = getattr(e, "summary", "") or getattr(e, "description", "")
                    link  = getattr(e, "link", "")
                    pub   = getattr(e, "published", "") or getattr(e, "updated", "")
                    if title:
                        fecha_obj = parsear_fecha(pub)
                        items.append({
                            "title": title, "summary": summ,
                            "link": link, "region": nombre,
                            "fecha_pub": str(fecha_obj),
                            "fecha_obj": fecha_obj,
                        })
                s = stats.get(nombre, {"total": 0, "util": 0})
                s["total"] = s.get("total", 0) + len(items)
                stats[nombre] = s
                return items, nombre
            except Exception:
                return [], nombre

        # ── Fase 1: feeds RSS ──
        _upd("Fase 1 / 3 — Consultando feeds RSS...",
             f"0 / {total_feeds} feeds procesados")
        with ThreadPoolExecutor(max_workers=len(FEEDS)) as ex:
            futures = {ex.submit(fetch, n, u): (n, u) for n, u in FEEDS}
            for fut in as_completed(futures):
                items, nombre = fut.result()
                todos.extend(items)
                feeds_ok[0] += 1
                _upd(f"Fase 1 / 3 — Consultando feeds RSS...",
                     f"{feeds_ok[0]} / {total_feeds} feeds  ·  {len(todos)} artículos")

        todos = deduplicar(todos)
        _upd(f"Fase 2 / 3 — Clasificando {len(todos)} artículos únicos...",
             "Aplicando keywords + MITRE ATT&CK")

        # ── Fase 2: clasificar + extraer IoC ──
        relevantes = []
        for art in todos:
            cat, metodo = clasificar_cyber(art["title"], art["summary"])
            if cat:
                art["_cat"] = cat
                art["_metodo"] = "kw"
                relevantes.append(art)
            elif metodo == "revisar":
                art["_cat"] = None
                art["_metodo"] = "ollama"
                relevantes.append(art)

        kw_arts     = [a for a in relevantes if a["_metodo"] == "kw"]
        ollama_arts = [a for a in relevantes if a["_metodo"] == "ollama"]

        _upd(f"Fase 3 / 3 — Descargando artículos para extracción de IoC...",
             f"{len(kw_arts)} por keywords  ·  {len(ollama_arts)} para Ollama")

        resultados  = []
        kw_count    = 0
        ollama_count = 0
        procesados  = [0]

        def procesar_kw(art):
            n = self._construir(art, art["_cat"])
            procesados[0] += 1
            if procesados[0] % 5 == 0:
                _upd(f"Fase 3 / 3 — Descargando artículos para extracción de IoC...",
                     f"{procesados[0]} / {len(relevantes)} procesados")
            return n

        def procesar_ollama(art):
            res = ollama_analizar_cyber(art["title"], art["summary"])
            procesados[0] += 1
            if procesados[0] % 5 == 0:
                _upd(f"Fase 3 / 3 — Descargando artículos + Ollama...",
                     f"{procesados[0]} / {len(relevantes)} procesados")
            if res and res.get("categoria") not in ("IRRELEVANTE", None):
                return self._construir_ollama(art, res)
            return None

        # Procesar en paralelo
        with ThreadPoolExecutor(max_workers=24) as ex:
            futs_kw = [ex.submit(procesar_kw, a) for a in kw_arts]
            futs_ol = [ex.submit(procesar_ollama, a) for a in ollama_arts]
            for f in as_completed(futs_kw):
                r = f.result()
                if r:
                    resultados.append(r)
                    kw_count += 1
            for f in as_completed(futs_ol):
                r = f.result()
                if r:
                    resultados.append(r)
                    ollama_count += 1

        stats_guardar(stats)
        self.after(0, lambda: self._poblar(resultados, kw_count, ollama_count))

    def _empaquetar(self, art, cat, ioc, mi_id, mi_name, mi_all, sev, resumen,
                    actor_extra=None, victima_extra=None,
                    pais_v_extra=None, cve_extra=None):
        """Construye el dict de noticia unificado."""
        grupos_str = ", ".join(ioc["grupos"]) or actor_extra or ""
        return {
            "cat": cat, "title": art["title"], "resumen": resumen,
            "link": art["link"], "region": art["region"],
            "fecha_pub": art.get("fecha_pub", str(date.today())),
            "fecha_obj": art.get("fecha_obj", date.today()),
            "mitre_id": mi_id, "mitre_name": mi_name, "mitre_all": mi_all,
            # ── contexto ──
            "actor":            grupos_str,
            "victima":          ioc.get("victima") or victima_extra or "",
            "sector":           ioc.get("sector_victima") or "",
            "pais_victima":     ioc.get("pais_victima") or pais_v_extra or "",
            "pais_origen":      ioc.get("pais_origen") or "",
            "impacto":          ioc.get("impacto_resumen") or "",
            "datos_robados":    ioc.get("datos_robados") or "",
            "sistemas_afect":   ioc.get("sistemas_afectados") or "",
            "software":         ", ".join(ioc.get("software_afectado", [])),
            "versiones":        ", ".join(ioc.get("versiones", [])),
            # ── indicadores técnicos ──
            "cve":      cve_extra or ", ".join(ioc["cve"]),
            "cvss":     ioc.get("cvss_score") or "",
            "tecnicas": ", ".join(ioc.get("tecnicas", [])),
            "ips":      ", ".join(ioc.get("ips", [])[:6]),
            "puertos":  ", ".join(ioc.get("puertos", [])[:5]),
            "dominios": ", ".join(ioc.get("dominios_mal", [])[:5]),
            "hashes":   ", ".join((ioc["hashes_md5"] + ioc["hashes_sha"])[:4]),
            "wallets":  ", ".join(
                ioc.get("btc_wallets",[]) + ioc.get("eth_wallets",[]) +
                ioc.get("xmr_wallets",[]))[:80],
            "emails":   ", ".join(ioc.get("emails",[])[:4]),
            "rescate":  ioc.get("rescate") or "",
            "severidad": sev, "incluir": False,
        }

    def _construir(self, art, cat):
        summ = art["summary"]
        texto_full = "" if len(summ) > 600 else _fetch_articulo(art.get("link", ""))
        texto_base = texto_full or summ
        ioc    = extraer_ioc(art["title"], summ, art.get("link", ""), texto_full)
        mitre  = detectar_mitre(art["title"], summ + " " + texto_base[:1000])
        mi_id, mi_name, mi_all = mitre if mitre else (None, None, [])
        sev    = nivel_severidad(cat, mi_all, ioc)
        resumen = condensar_resumen(art["title"], texto_base[:800])
        return self._empaquetar(art, cat, ioc, mi_id, mi_name, mi_all, sev, resumen)

    def _construir_ollama(self, art, res):
        cat = res.get("categoria", "CIBERCRIMEN / DARKWEB")
        if cat not in CATEGORIAS:
            cat = "CIBERCRIMEN / DARKWEB"
        summ = art["summary"]
        texto_full = "" if len(summ) > 600 else _fetch_articulo(art.get("link", ""))
        texto_base = texto_full or summ
        ioc     = extraer_ioc(art["title"], summ, art.get("link", ""), texto_full)
        mi_id   = res.get("mitre_tactica")
        mi_name = MITRE_TACTICS.get(mi_id, "") if mi_id else ""
        mi_all  = [mi_id] if mi_id else []
        sev     = nivel_severidad(cat, mi_all, ioc)
        resumen = res.get("resumen_es") or condensar_resumen(art["title"], texto_base[:800])
        return self._empaquetar(
            art, cat, ioc, mi_id, mi_name, mi_all, sev, resumen,
            actor_extra=res.get("actor"),
            victima_extra=res.get("victima"),
            pais_v_extra=res.get("pais_victima"),
            cve_extra=res.get("cve"),
        )

    def _poblar(self, resultados, kw_count, ollama_count):
        self._ocultar_carga()
        self._buscando = False

        # Ordenar: más nuevo primero por defecto
        resultados.sort(key=lambda n: n.get("fecha_obj", date.today()), reverse=True)

        for n in resultados:
            idx = next(self._id_counter)
            self.noticias_data[idx] = n
            mitre_txt = ""
            if n.get("mitre_id"):
                mitre_txt = f"{n['mitre_id']} · {n.get('mitre_name','')}"
            sev     = n.get("severidad", "BAJO")
            cat     = n["cat"]
            cat_safe = cat.replace(" ", "_").replace("/", "_")
            fecha   = n.get("fecha_pub", "")[:10]
            iid = self.tree.insert("", "end",
                values=(fecha, sev, cat, mitre_txt,
                        n["title"][:85],
                        (n.get("actor") or "—")[:30],
                        n["region"][:22]),
                tags=(cat_safe, f"SEV_{sev}"))
            self.noticias_vars[iid] = idx

        total = len(resultados)
        criticos = sum(1 for n in resultados if n.get("severidad") == "CRÍTICO")
        altos    = sum(1 for n in resultados if n.get("severidad") == "ALTO")
        self._lbl_count.config(text=f"  {total} amenazas detectadas")
        msg = f"✅  {total} amenazas  ·  kw: {kw_count}  ·  Ollama: {ollama_count}"
        self._lbl_status_footer.config(text=msg, fg=C["green"])
        right_msg = f"🔴 CRÍTICO: {criticos}  ·  🟠 ALTO: {altos}  ·  Total: {total}"
        self._lbl_status_right.config(text=right_msg, fg=C["txt"])

    # ─── DETALLE ───
    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[-1]
        idx = self.noticias_vars.get(iid)
        if idx is None:
            return
        n = self.noticias_data.get(idx, {})

        sev = n.get("severidad", "BAJO")
        sev_fg, sev_bg = SEVERITY_COLORS.get(sev, (C["txt"], C["bg1"]))
        self._lbl_sev.config(text=f"◈ {sev}", fg=sev_fg, bg=C["bg1"])

        cat = n.get("cat", "")
        _, cat_fg = CAT_COLORS.get(cat, (C["bg2"], C["cyan"]))
        self._lbl_cat.config(text=f"▸ {cat}", fg=cat_fg, bg=C["bg1"])

        mitre_txt = "Sin táctica MITRE detectada"
        if n.get("mitre_id"):
            mitre_txt = f"{n['mitre_id']}  ·  {n.get('mitre_name','')}"
            others = n.get("mitre_all", [])
            if len(others) > 1:
                extra = [f"{m} ({MITRE_TACTICS.get(m,'')})" for m in others[1:4]]
                mitre_txt += "\n+ " + "  |  ".join(extra)
        if n.get("tecnicas"):
            mitre_txt += f"\nT: {n['tecnicas']}"
        self._lbl_mitre.config(text=mitre_txt)

        def _ioc(key, val):
            lbl = self._ioc_labels.get(key)
            if lbl:
                lbl.config(text=val if val else "—")

        # ── Contexto ──
        _ioc("actor",    n.get("actor"))
        _ioc("pais_o",   n.get("pais_origen"))
        _ioc("victima",  n.get("victima"))
        _ioc("sector",   n.get("sector"))
        _ioc("pais_v",   n.get("pais_victima"))
        _ioc("impacto",  n.get("impacto"))
        _ioc("datos",    n.get("datos_robados"))
        _ioc("sistemas", n.get("sistemas_afect"))
        _ioc("rescate",  n.get("rescate"))
        # ── Técnicos ──
        _ioc("cve",      n.get("cve"))
        cvss = n.get("cvss")
        _ioc("cvss",     f"{cvss} / 10.0" if cvss else None)
        _ioc("software", n.get("software"))
        _ioc("versiones",n.get("versiones"))
        _ioc("tecnicas", n.get("tecnicas"))
        _ioc("ips",      n.get("ips"))
        _ioc("puertos",  n.get("puertos"))
        _ioc("dominios", n.get("dominios"))
        _ioc("hashes",   n.get("hashes"))
        _ioc("wallets",  n.get("wallets"))
        _ioc("emails",   n.get("emails"))

        self._txt_detail.config(state="normal")
        self._txt_detail.delete("1.0", "end")
        titulo  = n.get("title", "")
        resumen = n.get("resumen", "")
        fecha   = n.get("fecha_pub", "")
        fuente  = n.get("region", "")
        self._txt_detail.insert("end", f"{titulo}\n\n")
        self._txt_detail.insert("end", f"Fecha: {fecha}  ·  Fuente: {fuente}\n\n")
        self._txt_detail.insert("end", resumen)
        self._txt_detail.config(state="disabled")

        self._link_actual = n.get("link", "")

    def _abrir_link_double(self, event):
        link = getattr(self, "_link_actual", "")
        if link:
            import webbrowser
            webbrowser.open(link)

    def _mostrar_info(self):
        import tkinter.messagebox as mb
        msg = (
            "RMSecurity · Threat Intelligence Tool\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "¿Cómo usar el programa?\n\n"
            "1. BUSCAR  →  Ingresá palabras clave (ej: ransomware, APT,\n"
            "   CVE-2024) y presioná ◉ BUSCAR NOTICIAS.\n\n"
            "2. EXPLORAR  →  Hacé clic en cualquier noticia para ver\n"
            "   el análisis completo en el panel derecho.\n"
            "   Doble-click abre el artículo en el navegador.\n\n"
            "3. SELECCIONAR  →  Usá clic + Ctrl o Shift para\n"
            "   seleccionar múltiples noticias.\n\n"
            "4. EXPORTAR  →  Con las noticias seleccionadas,\n"
            "   presioná ⬡ Guardar IoC para generar el Excel.\n\n"
            "5. EXCEL  →  Abre el reporte THREAT_INTEL_IOC.xlsx\n"
            "   con hoja de Resumen y datos por severidad/categoría.\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "Fuentes: 47 feeds RSS · MITRE ATT&CK\n"
            "Extracción: CVEs · IPs · Hashes · APT · CVSS\n"
            "Opcional: Ollama LLM para enriquecimiento semántico"
        )
        mb.showinfo("ℹ  Cómo usar RMSecurity", msg)

    def _copiar_ioc_field(self, widget, key):
        val = widget.cget("text")
        if val and val != "—":
            self.clipboard_clear()
            self.clipboard_append(val)
            orig_fg = widget.cget("fg")
            orig_bg = widget.cget("bg")
            widget.config(fg="#000000", bg=C["cyan"])
            self.after(350, lambda: widget.config(fg=orig_fg, bg=orig_bg))

    def _copiar_informe_rapido(self):
        sel = self.tree.selection()
        iid = sel[-1] if sel else None
        idx = self.noticias_vars.get(iid) if iid else None
        n   = self.noticias_data.get(idx) if idx is not None else None
        if not n:
            messagebox.showinfo("Sin selección", "Seleccioná una noticia primero.")
            return

        lines = []
        lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        lines.append("  INFORME DE AMENAZA · THREAT INTELLIGENCE")
        lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"  {n.get('title','')}")
        lines.append("")
        lines.append(f"  Fecha:      {n.get('fecha_pub','')[:10]}")
        lines.append(f"  Severidad:  {n.get('severidad','—')}")
        lines.append(f"  Categoría:  {n.get('cat','—')}")
        mitre = n.get('mitre_id','')
        mname = n.get('mitre_name','')
        if mitre:
            lines.append(f"  MITRE:      {mitre} · {mname}")
        lines.append("")
        lines.append("  ── ACTORES ──────────────────────")

        def _f(label, key):
            v = n.get(key)
            if v:
                lines.append(f"  {label:<14}{v}")

        _f("Actor/Grupo:",  "actor")
        _f("País origen:",  "pais_origen")
        _f("Víctima:",      "victima")
        _f("Sector:",       "sector")
        _f("País víctima:", "pais_victima")
        lines.append("")
        lines.append("  ── IMPACTO ──────────────────────")
        _f("Impacto:",      "impacto")
        _f("Datos robados:","datos_robados")
        _f("Sistemas:",     "sistemas_afect")
        _f("Rescate:",      "rescate")
        lines.append("")
        lines.append("  ── TÉCNICOS ─────────────────────")
        _f("CVE:",          "cve")
        _f("CVSS:",         "cvss")
        _f("Software:",     "software")
        _f("Versiones:",    "versiones")
        _f("Técnicas:",     "tecnicas")
        _f("IPs (C2):",     "ips")
        _f("Dominios:",     "dominios")
        _f("Hashes:",       "hashes")
        _f("Wallets:",      "wallets")
        lines.append("")
        lines.append("  ── RESUMEN ──────────────────────")
        resumen = n.get("resumen", "")
        for chunk in [resumen[i:i+70] for i in range(0, min(len(resumen), 420), 70)]:
            lines.append(f"  {chunk}")
        if n.get("link"):
            lines.append("")
            lines.append(f"  Fuente: {n.get('link','')}")
        lines.append("")
        lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        lines.append("  Generado por RMSecurity Threat Intelligence Tool")
        lines.append("  Desarrollado por Rodrigo Moses")
        lines.append("  linkedin.com/in/rodrigo-m-793b36152")
        lines.append("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

        texto = "\n".join(lines)
        self.clipboard_clear()
        self.clipboard_append(texto)

        self._btn_informe.config(text="✔ Copiado!", fg=C["green"])
        self.after(1800, lambda: self._btn_informe.config(text="📋 Informe", fg=C["amber"]))

    def _aplicar_filtro(self):
        cat_f = self._cat_filter.get()
        sev_f = self._sev_filter.get()
        txt_f = self._txt_search.get().lower().strip()
        todos = list(self.noticias_vars.keys())

        # Detectar si el texto es un nombre de país para resaltado
        paises_match = set()
        if txt_f:
            for iid in todos:
                n = self.noticias_data.get(self.noticias_vars.get(iid, -1), {})
                pv = (n.get("pais_victima") or "").lower()
                po = (n.get("pais_origen")  or "").lower()
                if txt_f in pv or txt_f in po:
                    paises_match.add(iid)

        visibles = 0
        for iid in todos:
            idx = self.noticias_vars.get(iid)
            if idx is None:
                continue
            n  = self.noticias_data[idx]
            ok = True
            if cat_f not in ("TODAS", "ALL") and n.get("cat") != cat_f:
                ok = False
            if sev_f not in ("TODAS", "ALL") and n.get("severidad") != sev_f:
                ok = False
            if txt_f:
                haystack = " ".join(filter(None, [
                    n.get("title"),    n.get("resumen"),
                    n.get("actor"),    n.get("victima"),
                    n.get("cve"),      n.get("region"),
                    n.get("ips"),      n.get("hashes"),
                    n.get("dominios"), n.get("wallets"),
                    n.get("tecnicas"), n.get("software"),
                    n.get("pais_victima"), n.get("pais_origen"),
                    n.get("sector"),   n.get("impacto"),
                    n.get("datos_robados"),
                ])).lower()
                if txt_f not in haystack:
                    ok = False
            try:
                if ok:
                    self.tree.reattach(iid, "", "end")
                    cat_safe = n["cat"].replace(" ","_").replace("/","_")
                    sev = n.get("severidad","BAJO")
                    # Resaltar en amarillo si el término coincide con un país
                    if iid in paises_match:
                        self.tree.item(iid, tags=("match_pais",))
                    else:
                        self.tree.item(iid, tags=(cat_safe, f"SEV_{sev}"))
                    visibles += 1
                else:
                    self.tree.detach(iid)
            except Exception:
                pass

        total = len(self.noticias_data)
        if txt_f or cat_f not in ("TODAS","ALL") or sev_f not in ("TODAS","ALL"):
            self._lbl_count.config(
                text=f"  {visibles} de {total} amenazas",
                fg=C["cyan"] if visibles < total else C["txt"])
        else:
            self._lbl_count.config(text=f"  {total} amenazas detectadas", fg=C["txt"])

        if self._sort_col == "fecha":
            self._ordenar("fecha")

    def _guardar_ioc(self):
        sel_iids = self.tree.selection()
        if not sel_iids:
            messagebox.showinfo("Sin selección",
                                "Seleccioná una o más noticias antes de exportar.\n"
                                "(Clic para seleccionar · Ctrl+clic para múltiples)")
            return
        marcadas = [self.noticias_data[self.noticias_vars[iid]]
                    for iid in sel_iids if iid in self.noticias_vars]
        if not marcadas:
            return
        self.lbl_status.config(text=f"⟳ Guardando IoC de {len(marcadas)} noticias...")
        def _run():
            try:
                nuevas = acumular_ioc(marcadas)
                msg = f"✅  +{nuevas} registros guardados → THREAT_INTEL_IOC.xlsx"
                self.after(0, lambda: self.lbl_status.config(text=msg))
                if os.path.exists(_ioc_path):
                    os.startfile(_ioc_path)
            except Exception as e:
                self.after(0, lambda err=str(e): self.lbl_status.config(
                    text=f"Error: {err}"))
        threading.Thread(target=_run, daemon=True).start()

    def _abrir_ioc(self):
        if os.path.exists(_ioc_path):
            os.startfile(_ioc_path)
        else:
            messagebox.showinfo("Sin datos",
                                "Todavía no hay datos exportados.\n"
                                "Seleccioná noticias y presioná ⬡ Guardar IoC.")

    def _abrir_link(self):
        link = getattr(self, "_link_actual", "")
        if link:
            import webbrowser
            webbrowser.open(link)

    def _limpiar(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.noticias_data.clear()
        self.noticias_vars.clear()
        self._lbl_count.config(text="")
        self._lbl_status_right.config(text="")
        self.lbl_status.config(text="Limpiado.")
        self._txt_detail.config(state="normal")
        self._txt_detail.delete("1.0", "end")
        self._txt_detail.config(state="disabled")
        for lbl in self._ioc_labels.values():
            lbl.config(text="—")

    def _ver_feeds(self):
        stats = stats_cargar()
        win = tk.Toplevel(self)
        win.title("Estadísticas de Feeds")
        win.geometry("520x560")
        win.configure(bg=C["bg0"])
        tk.Label(win, text="PRODUCTIVIDAD DE FEEDS", bg=C["bg0"],
                 fg=C["cyan"], font=("Consolas", 10, "bold")).pack(pady=8)
        tree = ttk.Treeview(win, columns=("feed", "total"),
                            show="headings")
        tree.heading("feed",  text="Feed")
        tree.heading("total", text="Artículos vistos")
        tree.column("feed",  width=340)
        tree.column("total", width=120, anchor="center")
        for nombre, url in FEEDS:
            s = stats.get(nombre, {})
            total = s.get("total", 0)
            tag = "v" if total > 10 else ("m" if total > 3 else "r")
            tree.insert("", "end", values=(nombre, total), tags=(tag,))
        tree.tag_configure("v", background="#051405", foreground=C["green"])
        tree.tag_configure("m", background="#100C00", foreground=C["yellow"])
        tree.tag_configure("r", background="#0F0004", foreground=C["red"])
        tree.pack(fill="both", expand=True, padx=8, pady=8)


if __name__ == "__main__":
    app = CyberIntelApp()
    app.mainloop()
